from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from urllib.parse import quote
import os

BASE = "/Users/bhavya242574/Library/CloudStorage/OneDrive-EXLService.com(I)Pvt.Ltd/Desktop/TAM/excel_data"
OUT = "/Users/bhavya242574/Library/CloudStorage/OneDrive-EXLService.com(I)Pvt.Ltd/Desktop/TAM/excel_data_inventory.xlsx"

FILES = {
    "F1": "20200803 PC-Strategy-Solution Development and GoToMarket PlanVer1.5.xlsx",
    "F2": "Captive Operations Details Jun 2021 v3 - Copy.xlsx",
    "F3": "Competitor Analysis PL and CL.xlsx",
    "F4": "EXL Insurance Competitors.xlsx",
    "F5": "P&C Competitor Analysis.xlsx",
}

VINTAGE = {
    "F1": "Aug 2020 (v1.5)",
    "F2": "Jun 2021 (v3)",
    "F3": "Undated (~2023+)",
    "F4": "2024+ (post-CoPilot)",
    "F5": "~Feb 2021",
}

def file_uri(alias):
    return "file://" + quote(os.path.join(BASE, FILES[alias]), safe="/")

def sheet_link(alias, sheet, label="Open"):
    return f'=HYPERLINK("{file_uri(alias)}#\'{sheet}\'!A1","{label}")'

# (#, file, sheet, rxc, per_row, category, captures, era, notes)
INV = [
    (1,  "F1", "Home", "24x14", "—", "Strategy", "Landing page / nav", "Aug 2020", ""),
    (2,  "F1", "Go-To-Market Plan", "75x3", "section", "Strategy", "GTM narrative + 10 packaged solutions (FNOL-aaS, PaaS, NB-PaaS, etc.)", "Aug 2020", ""),
    (3,  "F1", "Top Insurers and Brokers", "231x43", "company", "Clients", "224 insurers/brokers ranked by 2019 NWP + CxOs + InsureTech/TPA partners + EXL wallet share + renewal/TCV", "2019/20", ""),
    (4,  "F1", "Buyer Priorities", "109x49", "client + CxO + priority", "Clients", "C-suite VOC -> priority area -> EXL solution mapping", "Aug 2020", ""),
    (5,  "F1", "Competitor Analysis", "76x19", "engagement", "Competitors", "72 EXL-competitor engagements at insurer accounts", "Feb 2021", "Duplicate of F5"),
    (6,  "F1", "Dashboard", "140x22", "varies", "Strategy", "Multi-table summary view (shortlist + VOC roll-up)", "Aug 2020", "Multi-table sheet"),
    (7,  "F1", "Top Clients Shortlist", "33x31", "account", "Clients", "28 prioritized accounts + bucket strategy + CxO roster", "Aug 2020", ""),
    (8,  "F1", "Partnerships and Alliances", "87x8", "provider", "Partners", "Partner ratings by capability", "Aug 2020", ""),
    (9,  "F1", "Capability Gap Assessment ", "72x17", "solution-component", "Capabilities", "Gap matrix v1 (EXL capability vs. competition + partner-needed flags)", "Aug 2020", "v1 of 3 iterations (trailing space in sheet name)"),
    (10, "F1", "Capability Gap Assessment 2", "67x17", "solution-component", "Capabilities", "Gap matrix v2", "Aug 2020", "v2 of 3"),
    (11, "F1", "Capability Gap Assessment -Ver2", "67x19", "solution-component", "Capabilities", "Gap matrix v3 (adds numeric capability scale 1-5 + Solutions-ID)", "Aug 2020", "v3 - most evolved"),
    (12, "F1", "Build-Buy-Partner Analysis", "138x13", "varies", "Capabilities", "Narrative + Table A 'Solution -> Build / Buy / Partner' recommendations", "Aug 2020", "Multi-table sheet"),
    (13, "F1", "Collated Collaterals List", "47x6", "asset", "Strategy", "EXL marketing collateral inventory + URLs", "Aug 2020", ""),
    (14, "F1", "Reference Tab - Do not Delete", "42x12", "reference", "Reference", "Dropdown lookup ranges (no business data)", "—", "Do not modify"),
    (15, "F1", "Consolidated Solutions", "58x15", "solution-component", "Capabilities", "Solution catalog", "Aug 2020", "Near-duplicate of #20"),
    (16, "F1", "White Spaces - Experience", "23x13", "row label", "Capabilities", "Value-chain x Personal/Commercial/Specialty matrix", "Aug 2020", "Empty grid - scaffold only"),
    (17, "F1", "Insuretch Anlaysis - Roopak", "14x10", "gap", "Partners", "Gap -> partner mapping by value chain", "Aug 2020", ""),
    (18, "F1", "InsureTech Landscape", "88x29", "insuretech", "Partners", "InsureTech inventory + funding + 2019/2020 EXL rating + LOB tagging", "2019/20", ""),
    (19, "F1", "Partnership Landscape", "70x29", "provider", "Partners", "Broader partner inventory (InsureTech + Other) with solution-mapping flags", "Aug 2020", ""),
    (20, "F1", "P&C Solutions List", "65x15", "solution-component", "Capabilities", "Solution catalog", "Aug 2020", "Near-duplicate of #15"),
    (21, "F1", "Client Prioritization Framework", "22x7", "criterion", "Strategy", "Scoring framework: 5 criteria x 5 options (Combined Ratio, Outsource Y/N, TCV potential, etc.)", "Aug 2020", ""),
    (22, "F1", "Solution Evaluation Criteria", "1x1", "—", "Reference", "Empty (Home link only)", "—", "Stub"),
    (23, "F1", "Sheet3", "15x17", "account", "Clients", "Orphan: 14 named accounts with CxOs + GWP", "Aug 2020", "Orphan - no Home link"),
    (24, "F2", "Captives in India", "9x10", "captive", "Captives", "7 captives in India (Metlife, Allstate, Equitable, Guardian, Principal, AXA XL, Mercer) + FTEs + leader contacts + COVID-WFH split", "Jun 2021", "Two-row header (R1 question, R2 field)"),
    (25, "F2", "Captives in the Philippines", "9x10", "captive", "Captives", "7 captives in PH (Manulife, Sun Life, AIA Philamlife, FWD, InLife, AXA PH, QBE)", "Jun 2021", "Same shape as #24"),
    (26, "F3", "Personal Lines", "21x9", "function", "Competitors", "5 PL functions x scale / delivery locations / transformation philosophy / digital solutions / partners / showcase clients", "~2023+", "Looks like a WNS profile; below-table notes at rows 9-10"),
    (27, "F3", "Commercial", "21x9", "function", "Competitors", "4 CL functions, same shape", "~2023+", "Looks like a WNS profile"),
    (28, "F4", "P&C Clientwise", "54x5", "client-competitor pair", "Competitors", "Competitor at each EXL P&C client + FTE/$ + CoPilot Research notes", "2024+", "Two-row header band"),
    (29, "F4", "L&A Clientwise", "61x5", "client-competitor pair", "Competitors", "Same shape for L&A clients", "2024+", "Two-row header band"),
    (30, "F4", "Competitor wise", "47x3", "competitor-segment", "Competitors", "Inverted view: each competitor's capabilities across the insurance value chain", "2024+", ""),
    (31, "F5", "Competitor Analysis", "76x20", "engagement", "Competitors", "72 engagements", "Feb 2021", "Duplicate of #5; small legend table to the right"),
]

CAT_COLORS = {
    "Strategy":     "FFF2CC",
    "Clients":      "DEEBF7",
    "Competitors":  "FCE4D6",
    "Capabilities": "E2EFDA",
    "Partners":     "EDEDED",
    "Captives":     "FFE699",
    "Reference":    "D9D9D9",
}

wb = Workbook()

# ============ Sheet 1: Inventory ============
ws = wb.active
ws.title = "Inventory"

headers = ["#", "File", "Sheet", "Open", "R x C", "Per-row entity", "Category", "What it captures", "Era", "Notes"]
ws.append(headers)

hdr_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
hdr_fill = PatternFill("solid", start_color="1F4E78")
hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin = Side(border_style="thin", color="BFBFBF")
brd = Border(left=thin, right=thin, top=thin, bottom=thin)

for c, _ in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=c)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.alignment = hdr_align
    cell.border = brd

data_font = Font(name="Arial", size=10)
link_font = Font(name="Arial", size=10, color="0563C1", underline="single")
data_align = Alignment(vertical="top", wrap_text=True)

for i, (num, alias, sheet, rxc, per_row, cat, captures, era, notes) in enumerate(INV, start=2):
    vals = [num, alias, sheet, sheet_link(alias, sheet), rxc, per_row, cat, captures, era, notes]
    for c, v in enumerate(vals, start=1):
        cell = ws.cell(row=i, column=c, value=v)
        cell.font = link_font if c == 4 else data_font
        cell.alignment = data_align
        cell.border = brd
    # category color band
    if cat in CAT_COLORS:
        ws.cell(row=i, column=7).fill = PatternFill("solid", start_color=CAT_COLORS[cat])

widths = {"A": 4, "B": 6, "C": 35, "D": 7, "E": 8, "F": 22, "G": 13, "H": 65, "I": 12, "J": 32}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

ws.freeze_panes = "A2"
ws.auto_filter.ref = ws.dimensions
ws.row_dimensions[1].height = 30

# ============ Sheet 2: Files ============
fws = wb.create_sheet("Files")
fheaders = ["Alias", "File Name", "Full Path", "Open File", "Vintage"]
fws.append(fheaders)
for c, _ in enumerate(fheaders, start=1):
    cell = fws.cell(row=1, column=c)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.alignment = hdr_align
    cell.border = brd

for i, (alias, fname) in enumerate(FILES.items(), start=2):
    full = os.path.join(BASE, fname)
    link = f'=HYPERLINK("{file_uri(alias)}","Open")'
    vals = [alias, fname, full, link, VINTAGE[alias]]
    for c, v in enumerate(vals, start=1):
        cell = fws.cell(row=i, column=c, value=v)
        cell.font = link_font if c == 4 else data_font
        cell.alignment = data_align
        cell.border = brd

fws.freeze_panes = "A2"
for col, w in zip("ABCDE", [8, 65, 95, 11, 22]):
    fws.column_dimensions[col].width = w

# ============ Sheet 3: Legend ============
lws = wb.create_sheet("Legend")
lws.append(["Category", "Description"])
legend_rows = [
    ("Strategy",     "Narrative GTM thinking, frameworks, dashboards, marketing collateral"),
    ("Clients",      "Client / account intelligence - insurers, brokers, CxOs, priorities, wallet"),
    ("Competitors",  "Competitor engagement data - who serves whom, at what scale, with what capability"),
    ("Capabilities", "EXL solution catalogs, capability-gap analysis, white-space maps"),
    ("Partners",     "Partner ecosystem - InsureTechs, TPAs, alliances"),
    ("Captives",     "Insurance captive operations (India, Philippines)"),
    ("Reference",    "Lookup tables, frameworks, admin / stubs"),
]
for row in legend_rows:
    lws.append(list(row))

for c in range(1, 3):
    cell = lws.cell(row=1, column=c)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.alignment = hdr_align
    cell.border = brd

for r in range(2, 2 + len(legend_rows)):
    cat = lws.cell(row=r, column=1).value
    if cat in CAT_COLORS:
        lws.cell(row=r, column=1).fill = PatternFill("solid", start_color=CAT_COLORS[cat])
    for c in range(1, 3):
        cell = lws.cell(row=r, column=c)
        cell.font = data_font
        cell.alignment = data_align
        cell.border = brd

lws.column_dimensions["A"].width = 16
lws.column_dimensions["B"].width = 85
lws.freeze_panes = "A2"

# ============ Sheet 4: Duplicates / oddities ============
dws = wb.create_sheet("Duplicates & Oddities")
dws.append(["Issue", "Where", "Note"])
issues = [
    ("Same dataset, two locations", "F1 'Competitor Analysis' (row #5) == whole of F5 (row #31)", "72-engagement table. Treat F5 as canonical or merge."),
    ("Three iterations of same matrix", "F1 'Capability Gap Assessment' / '... 2' / '...-Ver2' (rows #9-11)", "Ver2 is most evolved (adds 1-5 scale + Solutions-ID). Use as canonical."),
    ("Near-duplicate catalog", "F1 'Consolidated Solutions' (#15) and 'P&C Solutions List' (#20)", "Same modules with minor wrapper differences. Pick one as canonical."),
    ("Empty stub", "F1 'Solution Evaluation Criteria' (#22)", "Only contains Home link. Either remove or backfill."),
    ("Orphan tab", "F1 'Sheet3' (#23)", "No nav link. Looks pasted in. 14 accounts with CxOs + GWP."),
    ("Empty scaffold", "F1 'White Spaces - Experience' (#16)", "Value-chain x LOB grid built but never filled in."),
    ("Two-row header band", "F2 sheets, F4 'P&C Clientwise' / 'L&A Clientwise'", "R1 = source-group label, R2 = field name. Easy to misparse."),
    ("Data validation warnings", "F1 (workbook-level)", "openpyxl warns about unsupported Data Validation extensions on load."),
    ("Stale CxO names", "F1 'Buyer Priorities' (#4)", "Some CxO cells show #N/A; many CxO records dated Aug 2020 - likely turned over."),
    ("Vintage spread", "Files span Aug 2020 -> 2024+", "Newer files don't supersede older ones; no version-of-truth."),
]
for row in issues:
    dws.append(list(row))

for c in range(1, 4):
    cell = dws.cell(row=1, column=c)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.alignment = hdr_align
    cell.border = brd

for r in range(2, 2 + len(issues)):
    for c in range(1, 4):
        cell = dws.cell(row=r, column=c)
        cell.font = data_font
        cell.alignment = data_align
        cell.border = brd

dws.column_dimensions["A"].width = 28
dws.column_dimensions["B"].width = 50
dws.column_dimensions["C"].width = 70
dws.freeze_panes = "A2"

wb.save(OUT)
print(f"Saved: {OUT}")

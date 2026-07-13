"""
Stage 1+2: Ingestion + Entity Normalization (v2)
=================================================
Parses every sheet of every file. Each row in every sheet is either ingested
into a typed output table OR logged with a skip reason. The companion script
audit.py reconciles total rows = ingested + skipped per sheet.

Every output row carries:
  - source_file, source_sheet, source_row    (provenance)
  - file_vintage                              (date the file's data reflects)
  - captured_date                             (when the row was captured - V1: == file_vintage)
  - ingestion_date                            (when this pipeline ran)

Outputs (in pipeline/data/):
  Client + EXL relationship:
    - clients_top_insurers.csv, clients_sheet3.csv, clients_top_clients_shortlist.csv
  Competitor engagements:
    - engagements_competitor_analysis.csv (F1 + F5)
    - engagements_f4_clientwise.csv (F4 P&C + L&A)
  Buyer / VOC priorities:
    - priorities.csv
  Competitor capability profiles:
    - competitor_caps.csv (F4 Competitor wise)
    - competitor_profiles.csv (F3 PL + CL — WNS profile)
  Solution + capability catalogs:
    - solutions.csv (F1 P&C Solutions List)
    - capability_gaps.csv (F1 Capability Gap Assessment -Ver2)
    - insuretech_gap_analysis.csv (F1 Insuretch Anlaysis - Roopak)
  Partner ecosystem:
    - partners_alliances.csv (F1 Partnerships and Alliances)
    - insuretech_landscape.csv (F1 InsureTech Landscape)
    - partnership_landscape.csv (F1 Partnership Landscape)
    - build_buy_partner.csv (F1 Build-Buy-Partner Analysis - Table A)
  Captives:
    - captives.csv (F2 India + PH)
  Reference / framework:
    - prioritization_framework.csv (F1 Client Prioritization Framework)
  Audit:
    - rows_log.csv             (every row of every sheet with status + reason)
    - aliases.csv              (the alias map)
    - unresolved.csv           (names not matched - manual review)
  Report:
    - INGEST_REPORT.md
"""

import openpyxl
import os, csv, warnings, re
from datetime import date
from collections import Counter, defaultdict
warnings.filterwarnings('ignore')

from pathlib import Path
REPO = Path(__file__).resolve().parents[2]                      # repo root (code/pipeline/<script>.py)
BASE = str(REPO / "input_data" / "corpus")                      # source spreadsheets F1-F5
OUT  = str(REPO / "produced_data" / "pipeline" / "data")        # produced CSVs
REPORT = str(REPO / "produced_data" / "pipeline" / "INGEST_REPORT.md")

F1 = openpyxl.load_workbook(os.path.join(BASE, "20200803 PC-Strategy-Solution Development and GoToMarket PlanVer1.5.xlsx"), data_only=True)
F2 = openpyxl.load_workbook(os.path.join(BASE, "Captive Operations Details Jun 2021 v3 - Copy.xlsx"), data_only=True)
F3 = openpyxl.load_workbook(os.path.join(BASE, "Competitor Analysis PL and CL.xlsx"), data_only=True)
F4 = openpyxl.load_workbook(os.path.join(BASE, "EXL Insurance Competitors.xlsx"), data_only=True)
F5 = openpyxl.load_workbook(os.path.join(BASE, "P&C Competitor Analysis.xlsx"), data_only=True)

WORKBOOKS = {"F1": F1, "F2": F2, "F3": F3, "F4": F4, "F5": F5}

# Best-effort dating of each file's data
FILE_VINTAGES = {
    "F1": "2020-08-03",  # F1 Home sheet shows "Last Updated 2020-08-03"
    "F2": "2021-06-30",  # Filename: "Jun 2021"
    "F3": "2023-01-01",  # Estimated; content references Skense, Mosaic, etc.
    "F4": "2024-06-01",  # Estimated; references "CoPilot Research"
    "F5": "2021-02-21",  # "2/21" appears as a date marker in column R
}

INGESTION_DATE = date.today().isoformat()

# Sheets that are intentionally skipped (no data ingested) with documented reason
EXPLICITLY_SKIPPED = {
    ("F1", "Home"): "Navigation / landing page",
    ("F1", "Go-To-Market Plan"): "Narrative strategy text in col B; no row-level structured data",
    ("F1", "Dashboard"): "Multi-table summary view derived from other tables; no source data",
    ("F1", "Capability Gap Assessment "): "Superseded by 'Capability Gap Assessment -Ver2' (canonical)",
    ("F1", "Capability Gap Assessment 2"): "Superseded by 'Capability Gap Assessment -Ver2' (canonical)",
    ("F1", "Collated Collaterals List"): "Marketing-collateral URL list; not lead-relevant data",
    ("F1", "Reference Tab - Do not Delete"): "Lookup ranges; reference data only",
    ("F1", "Consolidated Solutions"): "Near-duplicate of 'P&C Solutions List' (canonical)",
    ("F1", "White Spaces - Experience"): "Empty scaffold grid (no data filled in)",
    ("F1", "Solution Evaluation Criteria"): "Empty stub (only Home link, 1x1)",
}

# ============ Row tracking ============
# Every row of every sheet we touch gets logged here for audit reconciliation
rows_log = []  # list of (file, sheet, row, status, reason)

def log_row(file_alias, sheet_name, row, status, reason=""):
    rows_log.append((file_alias, sheet_name, row, status, reason))

def row_is_empty(ws, r, max_col=None):
    """True if every cell in the row is None/empty."""
    cols = max_col or ws.max_column
    for c in range(1, cols + 1):
        if ws.cell(r, c).value not in (None, ""):
            return False
    return True

# ============ Alias maps ============

CLIENT_ALIASES_RAW = [
    ("Travelers Group", ["travelers", "travelers companies inc", "travelers companies inc.", "travelers companies", "travelers group", "trv"]),
    ("Allstate", ["allstate", "allstate corp", "allstate corp.", "allstate ins group", "allstate insurance"]),
    ("Hartford", ["hartford", "hartford financial services", "the hartford"]),
    ("Hanover Insurance Group", ["hanover", "hanover insurance group inc", "hanover insurance group inc.", "hanover insurance group"]),
    ("Liberty Mutual", ["liberty mutual", "liberty mutual ins cos", "liberty mutual group", "liberty mutual insurance"]),
    ("Nationwide", ["nationwide", "nationwide mutual group", "nationwide insurance", "nationwide group"]),
    ("AIG", ["aig", "american international group"]),
    ("Zurich", ["zurich", "zurich na", "zurich insurance group", "zurich insurance"]),
    ("AON", ["aon", "aon (netherlands)"]),
    ("IAG Australia", ["iah australia", "iag australia"]),
    ("IAG NZ", ["iag nz"]),
    ("IAG", ["iag"]),
    ("QBE", ["qbe", "qbe insurance group ltd", "qbe insurance group ltd.", "qbe insurance"]),
    ("Aviva", ["aviva", "aviva uk"]),
    ("AmTrust", ["amtrust", "amtrust financial services"]),
    ("MAPFRE", ["mapfre"]),
    ("Hiscox US", ["hiscox us"]),
    ("Hiscox UK", ["hiscox uk", "hiscox uk (5 fte’s)"]),
    ("Hiscox", ["hiscox"]),
    ("USAA", ["usaa", "usaa group"]),
    ("State Farm", ["state farm", "state farm group", "state farm mutual automobile ins.", "state farm mutual automobile ins"]),
    ("Progressive", ["progressive", "progressive corp.", "progressive corp", "progressive ins group"]),
    ("Chubb", ["chubb", "chubb ina group"]),
    ("Berkshire Hathaway", ["berkshire hathaway", "berkshire hathaway ins"]),
    ("CNA", ["cna", "cna ins cos", "c-n-a", "c na"]),
    ("Swiss Re", ["swiss re", "swiss re ltd.", "swiss re ltd"]),
    ("Genworth", ["genworth", "genworth financial inc", "genworth financial inc."]),
    ("Suncorp", ["suncorp", "sun corp"]),
    ("Marsh", ["marsh", "marsh broker"]),
    ("MMC", ["mmc"]),
    ("One Call Care Management", ["one call care management", "one call care management (occm)", "occm"]),
    ("Utica National Insurance Group", ["utica national insurance group"]),
    ("TRYG", ["tryg", "tryg (nordics)"]),
    ("Optum", ["optum"]),
    ("RSA Group UK", ["rsa", "rsa group uk"]),
    ("BUPA UK", ["bupa uk", "bupa"]),
    ("Aspen", ["aspen"]),
    ("Root Insurance", ["root insurance", "root"]),
    ("AXA", ["axa", "axa xl", "xl catlin", "axa xl (xl catlin)"]),
    ("ING Group", ["ing group", "ing"]),
    ("Assurant", ["assurant"]),
    ("Everest", ["everest", "everest insurance"]),
    ("American Modern", ["american modern"]),
    ("Kemper", ["kemper"]),
    ("Mitchell Insurance", ["mitchell insurance"]),
    ("CCC Information Services", ["ccc information services"]),
    ("Canopius", ["canopious", "canopius"]),
    ("Prudential", ["prudential"]),
    ("Unum", ["unum"]),
    ("Transamerica", ["transamerica"]),
    ("Lincoln Financial", ["lincoln", "lincoln financial"]),
    ("Aflac", ["aflac"]),
    ("Primerica", ["primerica"]),
    ("Trustage", ["trustage"]),
    ("Pacific Life", ["pacific life"]),
    ("Prosperity Life", ["prosperity"]),
    ("Global Atlantic", ["global atlantic"]),
    ("Fortitude Re", ["fortitude re"]),
    ("NYL", ["nyl", "new york life"]),
    ("Equitable", ["equitable"]),
    ("MetLife", ["metlife"]),
    ("Mercer", ["mercer"]),
    ("Guardian Life", ["guardian", "guardian life insurance co", "guardian life"]),
    ("Principal Financial", ["principal"]),
    ("Manulife", ["manulife", "manulife data services inc.", "manulife data services inc"]),
    ("Sun Life", ["sun life", "sun life of canada"]),
    ("AIA Philamlife", ["aia philamlife"]),
    ("FWD Insurance", ["fwd insurance", "fwd"]),
    ("InLife", ["inlife", "insular life", "inlife (aka insular life)"]),
    ("AXA Philippines", ["axa philippines"]),
]

COMPETITOR_ALIASES_RAW = [
    ("Cognizant", ["cts", "cognizant", "cognizant technology solutions"]),
    ("TCS", ["tcs", "tata consultancy services"]),
    ("WNS", ["wns"]),
    ("Genpact", ["genpact"]),
    ("Sutherland", ["sutherland"]),
    ("Accenture", ["accenture"]),
    ("Capgemini", ["capgemini"]),
    ("DXC Technology", ["dxc", "dxc technology"]),
    ("Infosys", ["infosys"]),
    ("Wipro", ["wipro"]),
    ("NTT Data", ["ntt", "ntt data"]),
    ("Mphasis", ["mphasis"]),
    ("IBM", ["ibm"]),
    ("EY", ["ey", "ernst & young"]),
    ("Deloitte", ["deloitte"]),
    ("KPMG", ["kpmg"]),
    ("PwC", ["pwc"]),
    ("McKinsey", ["mckinsey"]),
    ("Teleperformance", ["teleperformance"]),
    ("Foundever", ["foundever", "sitel"]),
    ("Qualfon", ["qualfon"]),
    ("Verisk", ["verisk"]),
    ("Vertafore", ["vertafore"]),
    ("Insurity", ["insurity"]),
    ("Medata", ["medata"]),
    ("Zinnia", ["zinnia"]),
    ("SE2", ["se2"]),
    ("Coforge", ["coforge"]),
    ("ResourcePro", ["resourcepro", "resource pro"]),
    ("Quantiphi", ["quantiphi"]),
    ("Tiger Analytics", ["tiger analytics"]),
    ("Virtusa", ["virtusa"]),
    ("Xceedance", ["xceedance"]),
    ("Workday", ["workday"]),
    ("ADP", ["adp"]),
    ("UKG", ["ukg"]),
    ("Five9", ["five9"]),
    ("Verint", ["verint"]),
    ("Mendix", ["mendix"]),
    ("Congruent", ["congruent"]),
    ("Indra", ["indra"]),
    ("CGI", ["cgi"]),
    ("Interactions", ["interactions"]),
    ("LYNX", ["lynx"]),
    ("Microsoft", ["microsoft"]),
    ("Google", ["google"]),
    ("Atos", ["atos"]),
    ("Oliver Wyman", ["oliverwyman", "oliver wyman"]),
    ("Assurant GCC", ["assurant gcc"]),
    ("Corvel", ["corvel"]),
    ("Ivans", ["ivans"]),
    ("USHUR", ["ushur"]),
    ("Captive", ["captive"]),
]

def build_alias_lookup(raw):
    return {v.strip().lower(): canon for canon, variants in raw for v in variants}

CLIENT_LOOKUP = build_alias_lookup(CLIENT_ALIASES_RAW)
COMPETITOR_LOOKUP = build_alias_lookup(COMPETITOR_ALIASES_RAW)
unresolved_clients = Counter()
unresolved_competitors = Counter()

def norm_str(s):
    if s is None: return ""
    return re.sub(r'\s+', ' ', str(s).strip())

def canon_client(name):
    if not name: return None
    n = norm_str(name)
    if not n: return None
    key = n.lower().rstrip('.')
    if key in CLIENT_LOOKUP:
        return CLIENT_LOOKUP[key]
    for canon, variants in CLIENT_ALIASES_RAW:
        for v in variants:
            if v in key or key in v:
                if len(key) >= 4 and len(v) >= 4:
                    return canon
    unresolved_clients[n] += 1
    return n

def canon_competitor(name):
    if not name: return None
    n = norm_str(name)
    if not n: return None
    key = n.lower().rstrip('.')
    if key in COMPETITOR_LOOKUP:
        return COMPETITOR_LOOKUP[key]
    first_token = key.split()[0] if key else ""
    if first_token in COMPETITOR_LOOKUP:
        return COMPETITOR_LOOKUP[first_token]
    unresolved_competitors[n] += 1
    return n

def write_csv(filename, headers, rows):
    path = os.path.join(OUT, filename)
    with open(path, 'w', newline='', encoding='utf-8') as f:
        w = csv.writer(f)
        w.writerow(headers)
        w.writerows(rows)
    return path, len(rows)

def provenance(file_alias, sheet_name, row_num):
    """Returns the 6 provenance + datetime columns appended to every output row."""
    return [
        file_alias, sheet_name, row_num,
        FILE_VINTAGES.get(file_alias, ""),
        FILE_VINTAGES.get(file_alias, ""),  # captured_date = file_vintage in V1
        INGESTION_DATE,
    ]

PROV_HEADERS = ["source_file", "source_sheet", "source_row", "file_vintage", "captured_date", "ingestion_date"]

# ============ Helpers ============

def iter_sheet_with_log(file_alias, sheet_name, ws, header_rows, ingest_fn):
    """
    Walks all rows of a sheet, calls ingest_fn(r) for each data row.
    ingest_fn returns either an output row (list) to emit, or None to skip.
    Header rows (1..N) are logged as 'header'.
    Empty rows are logged as 'empty'.
    Rows ingest_fn returns None for are logged as 'skipped_non_data'.
    Emitted rows are logged as 'ingested'.
    Returns the list of emitted rows.
    """
    emitted = []
    for r in range(1, ws.max_row + 1):
        if r in header_rows:
            log_row(file_alias, sheet_name, r, "header", f"row in header rows {min(header_rows)}-{max(header_rows)}")
            continue
        if row_is_empty(ws, r):
            log_row(file_alias, sheet_name, r, "empty", "")
            continue
        out = ingest_fn(r)
        if out is not None:
            log_row(file_alias, sheet_name, r, "ingested", "")
            emitted.append(out + provenance(file_alias, sheet_name, r))
        else:
            log_row(file_alias, sheet_name, r, "skipped_non_data", "row has content but not in expected data shape")
    return emitted


# ============ Ingest sheets ============

# ----- F1: Top Insurers and Brokers (header rows 1-6) -----
def ingest_top_insurers():
    ws = F1['Top Insurers and Brokers']
    def fn(r):
        raw_name = ws.cell(r, 2).value
        if not raw_name: return None
        return [
            canon_client(raw_name), raw_name,
            ws.cell(r, 3).value, ws.cell(r, 4).value, ws.cell(r, 5).value,
            ws.cell(r, 6).value, ws.cell(r, 8).value, ws.cell(r, 9).value,
            ws.cell(r, 10).value, ws.cell(r, 11).value, ws.cell(r, 1).value,
            ws.cell(r, 12).value, ws.cell(r, 13).value, ws.cell(r, 14).value,
            ws.cell(r, 15).value, ws.cell(r, 16).value, ws.cell(r, 17).value,
            ws.cell(r, 18).value, ws.cell(r, 19).value, ws.cell(r, 20).value,
            ws.cell(r, 21).value, ws.cell(r, 22).value,
            ws.cell(r, 35).value, ws.cell(r, 37).value, ws.cell(r, 38).value,
            ws.cell(r, 41).value, ws.cell(r, 42).value, ws.cell(r, 43).value,
        ]
    return iter_sheet_with_log("F1", "Top Insurers and Brokers", ws, set(range(1, 7)), fn)

# ----- F1: Sheet3 (header row 1) -----
def ingest_sheet3():
    ws = F1['Sheet3']
    def fn(r):
        raw_name = ws.cell(r, 1).value
        if not raw_name: return None
        return [
            canon_client(raw_name), raw_name,
            ws.cell(r, 2).value, ws.cell(r, 3).value, ws.cell(r, 4).value,
            ws.cell(r, 5).value, ws.cell(r, 6).value, ws.cell(r, 7).value,
            ws.cell(r, 8).value, ws.cell(r, 9).value, ws.cell(r, 10).value,
            ws.cell(r, 11).value, ws.cell(r, 12).value, ws.cell(r, 13).value,
            ws.cell(r, 14).value, ws.cell(r, 15).value, ws.cell(r, 16).value,
            ws.cell(r, 17).value,
        ]
    return iter_sheet_with_log("F1", "Sheet3", ws, {1}, fn)

# ----- F1: Top Clients Shortlist (header rows 1-5; data 6-33; bucket col forward-fill) -----
def ingest_top_clients_shortlist():
    ws = F1['Top Clients Shortlist']
    current_bucket = [None]  # cell so closure can mutate
    def fn(r):
        bucket = ws.cell(r, 1).value
        if bucket: current_bucket[0] = bucket
        raw_name = ws.cell(r, 3).value
        if not raw_name: return None
        return [
            canon_client(raw_name), raw_name,
            current_bucket[0], ws.cell(r, 2).value, ws.cell(r, 4).value,
            ws.cell(r, 5).value, ws.cell(r, 6).value, ws.cell(r, 7).value,
            ws.cell(r, 15).value, ws.cell(r, 16).value, ws.cell(r, 17).value,
        ]
    return iter_sheet_with_log("F1", "Top Clients Shortlist", ws, set(range(1, 6)), fn)

# ----- F1 + F5: Competitor Analysis (header rows 1-4) -----
def ingest_competitor_analysis(workbook, file_alias, sheet_name):
    ws = workbook[sheet_name]
    def fn(r):
        client_raw = ws.cell(r, 2).value
        comp_raw = ws.cell(r, 6).value
        if not client_raw or not comp_raw: return None
        return [
            canon_client(client_raw), client_raw,
            canon_competitor(comp_raw), comp_raw,
            ws.cell(r, 3).value, ws.cell(r, 4).value, ws.cell(r, 5).value,
            ws.cell(r, 7).value,
            "Y" if ws.cell(r, 8).value == "Y" else "",
            "Y" if ws.cell(r, 9).value == "Y" else "",
            "Y" if ws.cell(r, 10).value == "Y" else "",
            "Y" if ws.cell(r, 11).value == "Y" else "",
            "Y" if ws.cell(r, 12).value == "Y" else "",
            ws.cell(r, 13).value, ws.cell(r, 14).value,
            ws.cell(r, 15).value, ws.cell(r, 16).value,
        ]
    return iter_sheet_with_log(file_alias, sheet_name, ws, set(range(1, 5)), fn)

# ----- F4: P&C Clientwise / L&A Clientwise (header rows 1-2; client col forward-fill) -----
def ingest_f4_clientwise(sheet_name, segment):
    ws = F4[sheet_name]
    current = [None, None]  # [canon, raw]
    def fn(r):
        client_raw = ws.cell(r, 1).value
        if client_raw:
            current[0] = canon_client(client_raw)
            current[1] = client_raw
        comp_raw = ws.cell(r, 2).value
        if not comp_raw or not current[0]: return None
        return [
            current[0], current[1],
            canon_competitor(comp_raw), comp_raw,
            segment, ws.cell(r, 3).value, ws.cell(r, 4).value, ws.cell(r, 5).value,
        ]
    return iter_sheet_with_log("F4", sheet_name, ws, set(range(1, 3)), fn)

# ----- F1: Buyer Priorities (header rows 1-6) -----
def ingest_buyer_priorities():
    ws = F1['Buyer Priorities']
    def fn(r):
        client_raw = ws.cell(r, 1).value
        if not client_raw: return None
        return [
            canon_client(client_raw), client_raw,
            ws.cell(r, 2).value, ws.cell(r, 3).value, ws.cell(r, 4).value,
            ws.cell(r, 5).value, ws.cell(r, 6).value, ws.cell(r, 7).value,
            ws.cell(r, 8).value, ws.cell(r, 9).value, ws.cell(r, 10).value,
            ws.cell(r, 11).value, ws.cell(r, 12).value, ws.cell(r, 13).value,
            ws.cell(r, 21).value,
        ]
    return iter_sheet_with_log("F1", "Buyer Priorities", ws, set(range(1, 7)), fn)

# ----- F4: Competitor wise (header row 1; competitor col forward-fill) -----
def ingest_f4_competitor_wise():
    ws = F4['Competitor wise']
    current = [None, None]
    def fn(r):
        comp_raw = ws.cell(r, 1).value
        if comp_raw:
            current[0] = canon_competitor(comp_raw)
            current[1] = comp_raw
        segment = ws.cell(r, 2).value
        if not segment or not current[0]: return None
        return [current[0], current[1], segment, ws.cell(r, 3).value]
    return iter_sheet_with_log("F4", "Competitor wise", ws, {1}, fn)

# ----- F1: Partnerships and Alliances (header rows 1-6; capability col forward-fill) -----
def ingest_partnerships_alliances():
    ws = F1['Partnerships and Alliances']
    current_cap = [None]
    def fn(r):
        cap = ws.cell(r, 3).value
        if cap: current_cap[0] = cap
        provider = ws.cell(r, 4).value
        if not provider or not current_cap[0]: return None
        return [current_cap[0], provider, ws.cell(r, 5).value]
    return iter_sheet_with_log("F1", "Partnerships and Alliances", ws, set(range(1, 7)), fn)

# ----- F1: Capability Gap Assessment -Ver2 (header rows 1-2; many cols need forward-fill) -----
def ingest_capability_gaps():
    ws = F1['Capability Gap Assessment -Ver2']
    cur = {"area": None, "solution": None, "sol_id": None, "component": None}
    def fn(r):
        if ws.cell(r, 2).value: cur["area"] = ws.cell(r, 2).value
        if ws.cell(r, 3).value: cur["solution"] = ws.cell(r, 3).value
        if ws.cell(r, 4).value: cur["sol_id"] = ws.cell(r, 4).value
        if ws.cell(r, 5).value: cur["component"] = ws.cell(r, 5).value
        feature = ws.cell(r, 6).value
        if not feature: return None  # need at least a feature to record
        return [
            cur["area"], cur["solution"], cur["sol_id"], cur["component"], feature,
            ws.cell(r, 7).value, ws.cell(r, 8).value, ws.cell(r, 9).value,
            ws.cell(r, 10).value,
            "Y" if ws.cell(r, 11).value == "Y" else "",
            "Y" if ws.cell(r, 12).value == "Y" else "",
            "Y" if ws.cell(r, 13).value == "Y" else "",
        ]
    return iter_sheet_with_log("F1", "Capability Gap Assessment -Ver2", ws, set(range(1, 3)), fn)

# ----- F1: Buyer Priorities col W/X solution-code lookup (post-pass) -----
# Rows where client column (A) is empty but cols W (23) + X (24) have solution-code/component data.
# These are an embedded reference table inside the Buyer Priorities sheet.
def ingest_solution_codes():
    ws = F1['Buyer Priorities']
    emitted = []
    # Only touch rows that haven't been ingested by ingest_buyer_priorities (i.e. client col A is empty)
    # But rows 7-20 ALSO have W/X data alongside client data - those are captured in priorities.csv's extra_solution_code col.
    # Here we capture standalone code/component pairs from rows 30+ where client is empty.
    for r in range(7, ws.max_row + 1):
        client_a = ws.cell(r, 1).value
        code = ws.cell(r, 23).value     # col W
        component = ws.cell(r, 24).value # col X
        if client_a:  # already captured by priorities ingest
            continue
        if not code or not component:
            continue
        # This row IS data we want; re-classify it from skipped_non_data to ingested
        # Find and update the existing log entry
        for i, entry in enumerate(rows_log):
            if entry[0] == "F1" and entry[1] == "Buyer Priorities" and entry[2] == r:
                rows_log[i] = ("F1", "Buyer Priorities", r, "ingested", "")
                break
        emitted.append([code, component] + provenance("F1", "Buyer Priorities", r))
    return emitted

# ----- F1: Build-Buy-Partner Analysis Table B (pivot table starting R57) -----
# Counts of providers by capability and EXL rating (High/Low/Medium).
def ingest_bbp_table_b():
    ws = F1['Build-Buy-Partner Analysis']
    emitted = []
    cur_capability = [None]
    for r in range(57, ws.max_row + 1):
        c3 = ws.cell(r, 3).value
        if not c3: continue
        # R57-58 are headers
        if c3 in ("Count of Providers", "Row Labels"): continue
        # Capability rows have c4/c5/c6 with non-empty values (counts) — they're capability headers
        # Provider rows have just one of c4/c5/c6 with a count
        # Distinguish: capability rows have ALL of c4/c5/c6 populated OR are short top-level names like "Advanced Analytics"
        c4 = ws.cell(r, 4).value
        c5 = ws.cell(r, 5).value
        c6 = ws.cell(r, 6).value
        # Heuristic: if c4 + c5 + c6 are all numeric, treat as capability-level summary; otherwise provider
        cap_summary = all(isinstance(v, (int, float)) for v in (c4, c5, c6))
        if cap_summary:
            cur_capability[0] = c3
            role = "capability_summary"
        else:
            role = "provider_row"
        # Find existing log entry; if status was skipped_non_data, upgrade to ingested
        for i, entry in enumerate(rows_log):
            if entry[0] == "F1" and entry[1] == "Build-Buy-Partner Analysis" and entry[2] == r:
                if entry[3] in ("skipped_non_data", "empty"):
                    rows_log[i] = ("F1", "Build-Buy-Partner Analysis", r, "ingested", "")
                break
        emitted.append([
            cur_capability[0], c3, role,
            c4 if isinstance(c4, (int, float)) else "",
            c5 if isinstance(c5, (int, float)) else "",
            c6 if isinstance(c6, (int, float)) else "",
        ] + provenance("F1", "Build-Buy-Partner Analysis", r))
    return emitted

# ----- F1: InsureTech Landscape (header rows 1-7; data 8+) -----
def ingest_insuretech_landscape():
    ws = F1['InsureTech Landscape']
    def fn(r):
        name = ws.cell(r, 2).value
        if not name: return None
        return [
            name,                    # InsureTech
            ws.cell(r, 3).value,     # What they do
            ws.cell(r, 4).value,     # Market
            ws.cell(r, 5).value,     # LOB
            ws.cell(r, 6).value,     # 2019 EXL Ranking
            ws.cell(r, 7).value,     # 2020 EXL Rating
            ws.cell(r, 8).value,     # Category
            ws.cell(r, 9).value,     # Value Chain
            ws.cell(r, 10).value,    # Market
            ws.cell(r, 11).value,    # Potential Funding
            ws.cell(r, 12).value,    # Vertical
            ws.cell(r, 13).value,    # Business Lines Supported
        ]
    return iter_sheet_with_log("F1", "InsureTech Landscape", ws, set(range(1, 8)), fn)

# ----- F1: Partnership Landscape (header rows 1-7) -----
def ingest_partnership_landscape():
    ws = F1['Partnership Landscape']
    def fn(r):
        name = ws.cell(r, 1).value
        if not name: return None
        return [
            name,                    # Provider
            ws.cell(r, 2).value,     # Type
            ws.cell(r, 3).value,     # What they do
            ws.cell(r, 4).value,     # Value Chain
            ws.cell(r, 5).value,     # Applicable Function
            ws.cell(r, 6).value,     # Market
            ws.cell(r, 7).value,     # LOB
            ws.cell(r, 8).value,     # Function/Capability
            ws.cell(r, 9).value,     # Inception Year
            ws.cell(r, 10).value,    # 2019 EXL Ranking
            ws.cell(r, 11).value,    # 2020 EXL Rating
            ws.cell(r, 12).value,    # No of solutions
        ]
    return iter_sheet_with_log("F1", "Partnership Landscape", ws, set(range(1, 8)), fn)

# ----- F1: P&C Solutions List (header rows 1-9; solution col forward-fill) -----
def ingest_solutions_list():
    ws = F1['P&C Solutions List']
    cur = {"num": None, "problem": None, "offering": None}
    def fn(r):
        if ws.cell(r, 1).value is not None: cur["num"] = ws.cell(r, 1).value
        if ws.cell(r, 2).value: cur["problem"] = ws.cell(r, 2).value
        if ws.cell(r, 3).value: cur["offering"] = ws.cell(r, 3).value
        module = ws.cell(r, 4).value
        # Some rows are "summary" rows with no module but with outcomes/metrics
        if not module and not ws.cell(r, 11).value:
            return None
        return [
            cur["num"], cur["offering"], module,
            ws.cell(r, 5).value,   # outcomes (only on first row of solution block)
            ws.cell(r, 6).value,   # metrics
            ws.cell(r, 7).value,   # buying center
            ws.cell(r, 8).value,   # client appetite / feedback
            ws.cell(r, 9).value,   # market appetite
            ws.cell(r, 10).value,  # existing capability in market
            ws.cell(r, 11).value,  # EXL Capabilities - Area
            ws.cell(r, 12).value,  # L/M/H
            ws.cell(r, 13).value,  # Details
        ]
    return iter_sheet_with_log("F1", "P&C Solutions List", ws, set(range(1, 10)), fn)

# ----- F1: Client Prioritization Framework (header rows 1-3; tiny reference table) -----
def ingest_prioritization_framework():
    ws = F1['Client Prioritization Framework']
    cur_table = [None]
    def fn(r):
        # Row 4 (and similar) = "Parameter | Option 1 | Option 2 ..." sub-header
        # Row 11 = second instance of header
        v_b = ws.cell(r, 2).value
        if v_b == "Parameter":
            cur_table[0] = r
            return None  # header
        if v_b and isinstance(v_b, str) and "Client Prioritization" in v_b:
            return None  # section title
        if not v_b: return None
        return [
            v_b,                    # Parameter
            ws.cell(r, 3).value,    # Option 1
            ws.cell(r, 4).value,    # Option 2
            ws.cell(r, 5).value,    # Option 3
            ws.cell(r, 6).value,    # Option 4
            ws.cell(r, 7).value,    # Option 5
        ]
    return iter_sheet_with_log("F1", "Client Prioritization Framework", ws, {1}, fn)

# ----- F1: Insuretch Anlaysis - Roopak (header rows 1-5) -----
def ingest_insuretech_roopak():
    ws = F1['Insuretch Anlaysis - Roopak']
    cur_chain = [None]
    def fn(r):
        chain = ws.cell(r, 2).value
        if chain: cur_chain[0] = chain
        func = ws.cell(r, 3).value
        if not func: return None
        return [
            cur_chain[0], func,
            ws.cell(r, 4).value,    # gap - ops capability
            ws.cell(r, 5).value,    # gap - digital solutions
            ws.cell(r, 6).value,    # gap - analytics
            ws.cell(r, 7).value,    # gap - platform
            ws.cell(r, 8).value,    # partner kind required
            ws.cell(r, 9).value,    # need for partner
            ws.cell(r, 10).value,   # potential partners
        ]
    return iter_sheet_with_log("F1", "Insuretch Anlaysis - Roopak", ws, set(range(1, 6)), fn)

# ----- F1: Build-Buy-Partner Analysis (narrative top + Table A; headers approx rows 1-13) -----
def ingest_build_buy_partner():
    ws = F1['Build-Buy-Partner Analysis']
    cur_sol = [None]
    def fn(r):
        # Data rows have a non-empty col B (solution name on first row of block, then component on subsequent)
        sol = ws.cell(r, 2).value
        component = ws.cell(r, 3).value
        if sol and (not component) and isinstance(sol, str) and sol.startswith("Solution"):
            cur_sol[0] = sol
            return None  # header-of-block row, no data yet
        if sol and isinstance(sol, str) and sol.startswith("Solution"):
            cur_sol[0] = sol
        if not component: return None
        return [
            cur_sol[0], component,
            ws.cell(r, 4).value,    # gap description
            ws.cell(r, 5).value,    # evaluation criteria
            ws.cell(r, 6).value,    # supplier applicability
            ws.cell(r, 7).value,    # InsureTech partner
            ws.cell(r, 8).value,    # TPA partner
            ws.cell(r, 9).value,    # Other partner
            ws.cell(r, 10).value,   # Potential to Build Y/N
            ws.cell(r, 11).value,   # Build Comments
        ]
    return iter_sheet_with_log("F1", "Build-Buy-Partner Analysis", ws, set(range(1, 14)), fn)

# ----- F2: Captives in India / Philippines (header rows 1-2) -----
def ingest_captives(sheet_name, country):
    ws = F2[sheet_name]
    def fn(r):
        name = ws.cell(r, 2).value
        if not name: return None
        return [
            canon_client(name), name, country,
            ws.cell(r, 3).value,    # location
            ws.cell(r, 4).value,    # FTEs
            ws.cell(r, 5).value,    # vertical/BU
            ws.cell(r, 6).value,    # areas of ops
            ws.cell(r, 7).value,    # type of ops
            ws.cell(r, 8).value,    # geos served
            ws.cell(r, 9).value,    # leader contact / subjective commentary
            ws.cell(r, 10).value,   # COVID impact
        ]
    return iter_sheet_with_log("F2", sheet_name, ws, set(range(1, 3)), fn)

# ----- F3: Personal Lines / Commercial (header rows 1-2; rows 9-10 = thematic notes) -----
def ingest_f3_competitor_profile(sheet_name, lob_segment):
    ws = F3[sheet_name]
    def fn(r):
        func = ws.cell(r, 3).value
        col_d = ws.cell(r, 4).value
        # Rows 9-10 are general thematic notes — col 3 is empty, col 4 has the note
        if not func and col_d:
            # Promote: capture as a thematic note
            return ["WNS", lob_segment, "(General Theme)", col_d, "", "", "", "", ""]
        if not func: return None
        return [
            "WNS", lob_segment, func,
            ws.cell(r, 4).value, ws.cell(r, 5).value, ws.cell(r, 6).value,
            ws.cell(r, 7).value, ws.cell(r, 8).value, ws.cell(r, 9).value,
        ]
    return iter_sheet_with_log("F3", sheet_name, ws, set(range(1, 3)), fn)


# ============ Run all ingestion ============

print("Running ingestion...")
top_insurers = ingest_top_insurers()
sheet3 = ingest_sheet3()
shortlist = ingest_top_clients_shortlist()
f1_comp = ingest_competitor_analysis(F1, "F1", "Competitor Analysis")
f5_comp = ingest_competitor_analysis(F5, "F5", "Competitor Analysis")
f4_pc = ingest_f4_clientwise("P&C Clientwise", "P&C")
f4_la = ingest_f4_clientwise("L&A Clientwise", "L&A")
priorities = ingest_buyer_priorities()
f4_comp = ingest_f4_competitor_wise()
partnerships = ingest_partnerships_alliances()
cap_gaps = ingest_capability_gaps()
insuretech_landscape = ingest_insuretech_landscape()
partnership_landscape = ingest_partnership_landscape()
solutions = ingest_solutions_list()
prio_framework = ingest_prioritization_framework()
insuretech_roopak = ingest_insuretech_roopak()
build_buy_partner = ingest_build_buy_partner()
captives_india = ingest_captives("Captives in India", "India")
captives_ph = ingest_captives("Captives in the Philippines", "Philippines")
f3_pl = ingest_f3_competitor_profile("Personal Lines", "Personal Lines")
f3_cl = ingest_f3_competitor_profile("Commercial", "Commercial")

# Post-passes (must run AFTER the main ingestions for those sheets):
solution_codes = ingest_solution_codes()
bbp_table_b = ingest_bbp_table_b()

# Log explicitly-skipped sheets
for (file_alias, sheet_name), reason in EXPLICITLY_SKIPPED.items():
    ws = WORKBOOKS[file_alias][sheet_name]
    for r in range(1, ws.max_row + 1):
        if row_is_empty(ws, r):
            log_row(file_alias, sheet_name, r, "empty", "")
        else:
            log_row(file_alias, sheet_name, r, "explicitly_skipped", reason)


# ============ Write output CSVs ============

write_csv("clients_top_insurers.csv",
    ["client_canonical", "name_raw", "cid", "type", "nwp_2019", "pct_change",
     "rank_2019", "geography", "combined_ratio", "relationship_type", "growth_leader",
     "ce", "ceo", "cfo", "cro", "cco", "coo", "cio_cto", "cuo", "outsource",
     "exl_competitors_freetext", "year_flag", "exl_client", "exl_share_of_wallet",
     "competition_account", "renewal_timeframe", "future_tcv", "client_priority_rating"]
    + PROV_HEADERS,
    top_insurers)

write_csv("clients_sheet3.csv",
    ["client_canonical", "name_raw", "geography", "strategic_named", "client_exec",
     "ceo", "cfo", "cro", "cco", "coo", "cio_cto", "cuo", "gwp", "growth_pct",
     "combined_ratio", "outsource", "exl_competitors", "year_flag"] + PROV_HEADERS,
    sheet3)

write_csv("clients_top_clients_shortlist.csv",
    ["client_canonical", "name_raw", "bucket_strategy", "priority_position",
     "geography", "strategic_named", "growth_leader", "client_exec",
     "gwp", "growth_pct", "combined_ratio"] + PROV_HEADERS,
    shortlist)

write_csv("engagements_competitor_analysis.csv",
    ["client_canonical", "client_raw", "competitor_canonical", "competitor_raw",
     "client_type", "strategic_named", "exl_contact", "contract_renewal_window",
     "claims_y", "uw_y", "premium_audit_y", "fna_y", "platform_y",
     "ftes", "comments", "current_status", "owner"] + PROV_HEADERS,
    f1_comp + f5_comp)

write_csv("engagements_f4_clientwise.csv",
    ["client_canonical", "client_raw", "competitor_canonical", "competitor_raw",
     "lob_segment", "ftes_or_revenue_freetext", "competitor_segment_supported",
     "research_notes"] + PROV_HEADERS,
    f4_pc + f4_la)

write_csv("priorities.csv",
    ["client_canonical", "client_raw", "cxo_designation", "cxo_name",
     "location", "exl_ce", "voc_source", "priority_area", "priority_sub_area",
     "opportunity_freetext", "sign_off_next_steps", "exl_solution",
     "exl_solution_component", "solution_life_cycle", "extra_solution_code"] + PROV_HEADERS,
    priorities)

write_csv("competitor_caps.csv",
    ["competitor_canonical", "competitor_raw", "value_chain_segment",
     "research_notes"] + PROV_HEADERS,
    f4_comp)

write_csv("partners_alliances.csv",
    ["function_capability", "provider", "exl_assessment_2020"] + PROV_HEADERS,
    partnerships)

write_csv("capability_gaps.csv",
    ["area", "solution", "solution_id", "component", "feature",
     "exl_capability", "exl_capability_scale_1_5", "challenges",
     "gap_description", "partner_insurtech", "partner_tpa", "partner_other"]
    + PROV_HEADERS,
    cap_gaps)

write_csv("insuretech_landscape.csv",
    ["insuretech", "what_they_do", "market_1", "lob", "ranking_2019", "rating_2020",
     "category", "value_chain", "market_2", "potential_funding", "vertical",
     "business_lines_supported"] + PROV_HEADERS,
    insuretech_landscape)

write_csv("partnership_landscape.csv",
    ["provider", "type", "what_they_do", "value_chain", "applicable_function",
     "market", "lob", "function_capability", "inception_year",
     "ranking_2019", "rating_2020", "num_solutions"] + PROV_HEADERS,
    partnership_landscape)

write_csv("solutions.csv",
    ["solution_num", "offering", "module", "outcomes", "metrics", "buying_center",
     "client_appetite", "market_appetite", "existing_capability_in_market",
     "exl_capabilities_area", "exl_capability_lmh", "details"] + PROV_HEADERS,
    solutions)

write_csv("prioritization_framework.csv",
    ["parameter", "option_1", "option_2", "option_3", "option_4", "option_5"]
    + PROV_HEADERS,
    prio_framework)

write_csv("insuretech_gap_analysis.csv",
    ["value_chain", "function", "gap_ops_capability", "gap_digital_solutions",
     "gap_analytics", "gap_platform", "partner_kind_required", "need_for_partner",
     "potential_partners"] + PROV_HEADERS,
    insuretech_roopak)

write_csv("build_buy_partner.csv",
    ["solution_name", "component", "gap_description", "evaluation_criteria",
     "supplier_applicability", "insuretech_partner", "tpa_partner", "other_partner",
     "potential_to_build_yn", "build_comments"] + PROV_HEADERS,
    build_buy_partner)

write_csv("captives.csv",
    ["client_canonical", "client_raw", "country", "location", "ftes",
     "vertical_bu", "areas_of_operations", "type_of_operations", "geos_served",
     "leader_subjective_commentary", "covid_impact"] + PROV_HEADERS,
    captives_india + captives_ph)

write_csv("competitor_profiles.csv",
    ["competitor", "lob_segment", "function", "scale", "delivery_locations",
     "transformation_philosophy", "digital_solutions", "partnerships",
     "showcase_clients"] + PROV_HEADERS,
    f3_pl + f3_cl)

write_csv("solution_codes.csv",
    ["solution_code", "solution_component"] + PROV_HEADERS,
    solution_codes)

write_csv("build_buy_partner_table_b.csv",
    ["capability", "provider_or_capability", "role", "high_count", "low_count", "medium_count"]
    + PROV_HEADERS,
    bbp_table_b)

# Audit log: every row of every sheet
write_csv("rows_log.csv",
    ["source_file", "source_sheet", "source_row", "status", "reason"],
    rows_log)

# Alias map
alias_rows = []
for canon, variants in CLIENT_ALIASES_RAW:
    for v in variants:
        alias_rows.append(["client", v, canon])
for canon, variants in COMPETITOR_ALIASES_RAW:
    for v in variants:
        alias_rows.append(["competitor", v, canon])
write_csv("aliases.csv", ["entity_kind", "variant", "canonical"], alias_rows)

# Unresolved
unresolved = []
for n, c in unresolved_clients.most_common():
    unresolved.append(["client", n, c])
for n, c in unresolved_competitors.most_common():
    unresolved.append(["competitor", n, c])
write_csv("unresolved.csv", ["entity_kind", "name_raw", "occurrences"], unresolved)


# ============ Report ============

# Stats from rows_log
status_counter = Counter(r[3] for r in rows_log)
sheets_touched = sorted({(r[0], r[1]) for r in rows_log})
total_rows = sum(len(WORKBOOKS[f][s].iter_rows() != None and [r for r in range(1, WORKBOOKS[f][s].max_row+1)]) for f, s in sheets_touched if False)  # simpler:
total_rows = 0
for wb_alias, wb in WORKBOOKS.items():
    for sn in wb.sheetnames:
        total_rows += wb[sn].max_row

ingested_rows = status_counter.get("ingested", 0)
header_rows = status_counter.get("header", 0)
empty_rows = status_counter.get("empty", 0)
skipped_rows = status_counter.get("explicitly_skipped", 0)
non_data_rows = status_counter.get("skipped_non_data", 0)

# Signal-relevant aggregates
clients_with_renewal_le_6mo = sum(1 for r in f5_comp if r[7] and "< 6 Months" in str(r[7]))
named_strategic = sum(1 for r in top_insurers if r[9] in ("Strategic", "Named", "Named Backup"))
outsourcing_yes = sum(1 for r in top_insurers if str(r[19]).lower().strip() == "yes")
cr_over_100 = sum(1 for r in top_insurers if isinstance(r[8], (int, float)) and r[8] > 100)
priorities_not_initiated = sum(1 for r in priorities if r[13] == "Not Initiated")

report_md = f"""# Ingestion Report

Generated: {os.popen('date').read().strip()}
Ingestion date: {INGESTION_DATE}

## File vintages (used as captured_date for every row)

| Alias | File | Inferred vintage | Source of date |
|---|---|---|---|
| F1 | 20200803 PC-Strategy-Solution Development and GoToMarket PlanVer1.5.xlsx | {FILE_VINTAGES['F1']} | Filename + Home sheet "Last Updated" |
| F2 | Captive Operations Details Jun 2021 v3 - Copy.xlsx | {FILE_VINTAGES['F2']} | Filename "Jun 2021" |
| F3 | Competitor Analysis PL and CL.xlsx | {FILE_VINTAGES['F3']} | **Estimated** - content references post-2022 tools (Skense, Mosaic) |
| F4 | EXL Insurance Competitors.xlsx | {FILE_VINTAGES['F4']} | **Estimated** - references "CoPilot Research" (post-2023) |
| F5 | P&C Competitor Analysis.xlsx | {FILE_VINTAGES['F5']} | "2/21" date marker in column R |

**Estimated vintages are best guesses** - flagged in `file_vintage` column on every row. Recommend confirming with file owners.

## Row accounting (every cell of every sheet)

| Status | Rows | Meaning |
|---|---:|---|
| ingested | {ingested_rows} | Produced an output row in one of the typed CSVs |
| header | {header_rows} | First N rows of a sheet, intentionally skipped |
| empty | {empty_rows} | All cells in the row are None/blank |
| explicitly_skipped | {skipped_rows} | Sheet flagged as not-ingested with documented reason |
| skipped_non_data | {non_data_rows} | Row has content but didn't fit the table's data shape (e.g. footer notes) |
| **TOTAL** | **{ingested_rows + header_rows + empty_rows + skipped_rows + non_data_rows}** | Should equal sum of max_row across all sheets ({total_rows}) |

Reconciliation: see `audit.py` for full per-sheet breakdown.

## Output tables

| Table | Rows | Source |
|---|---:|---|
| clients_top_insurers.csv | {len(top_insurers)} | F1 / Top Insurers and Brokers |
| clients_sheet3.csv | {len(sheet3)} | F1 / Sheet3 |
| clients_top_clients_shortlist.csv | {len(shortlist)} | F1 / Top Clients Shortlist |
| engagements_competitor_analysis.csv | {len(f1_comp) + len(f5_comp)} | F1 + F5 Competitor Analysis |
| engagements_f4_clientwise.csv | {len(f4_pc) + len(f4_la)} | F4 P&C + L&A Clientwise |
| priorities.csv | {len(priorities)} | F1 / Buyer Priorities |
| competitor_caps.csv | {len(f4_comp)} | F4 / Competitor wise |
| partners_alliances.csv | {len(partnerships)} | F1 / Partnerships and Alliances |
| capability_gaps.csv | {len(cap_gaps)} | F1 / Capability Gap Assessment -Ver2 |
| insuretech_landscape.csv | {len(insuretech_landscape)} | F1 / InsureTech Landscape |
| partnership_landscape.csv | {len(partnership_landscape)} | F1 / Partnership Landscape |
| solutions.csv | {len(solutions)} | F1 / P&C Solutions List |
| prioritization_framework.csv | {len(prio_framework)} | F1 / Client Prioritization Framework |
| insuretech_gap_analysis.csv | {len(insuretech_roopak)} | F1 / Insuretch Anlaysis - Roopak |
| build_buy_partner.csv | {len(build_buy_partner)} | F1 / Build-Buy-Partner Analysis |
| captives.csv | {len(captives_india) + len(captives_ph)} | F2 India + PH |
| competitor_profiles.csv | {len(f3_pl) + len(f3_cl)} | F3 PL + CL (WNS profile) |

## Sheets ingested

{chr(10).join('- ' + f + ' / ' + s for f, s in sorted(sheets_touched) if (f, s) not in EXPLICITLY_SKIPPED)}

## Sheets explicitly skipped (with reason)

{chr(10).join('- ' + f + ' / ' + s + ' — ' + r for (f, s), r in EXPLICITLY_SKIPPED.items())}

## Entity normalization

- Unique canonical clients: {len({r[0] for r in top_insurers + sheet3 + shortlist if r[0]})}
- Unique canonical competitors: {len({r[2] for r in (f1_comp + f5_comp + f4_pc + f4_la + f4_comp) if r[2]})}
- Client aliases: {sum(len(v) for _, v in CLIENT_ALIASES_RAW)} variants -> {len(CLIENT_ALIASES_RAW)} canonicals
- Competitor aliases: {sum(len(v) for _, v in COMPETITOR_ALIASES_RAW)} variants -> {len(COMPETITOR_ALIASES_RAW)} canonicals
- Unresolved client names: {len(unresolved_clients)} (mostly small carriers/brokers - see `unresolved.csv`)
- Unresolved competitor names: {len(unresolved_competitors)}

## Signal-relevant aggregates (for trigger-library design in Task #7)

| Signal | Count |
|---|---:|
| F5 engagements with renewal `< 6 Months` (canonical) | {clients_with_renewal_le_6mo} |
| Clients flagged Strategic / Named / Named Backup | {named_strategic} |
| Clients with Outsource = Yes | {outsourcing_yes} |
| Clients with Combined Ratio > 100 | {cr_over_100} |
| Priorities with status "Not Initiated" | {priorities_not_initiated} |

## Data quality flags

- **F5 is canonical for Competitor Analysis**: F1's Competitor Analysis tab duplicates F5's data with sparser Comments/Owner fields. Downstream consumers should use F5.
- **F1 Buyer Priorities is sparse**: only 3 clients (Travelers, Allstate, CNA) have VOC captured despite the sheet having 109 rows.
- **CxO names broken in Buyer Priorities**: rows beyond the first per client show `#N/A` (broken VLOOKUP).
- **Top Insurers col V header/content mismatch**: header reads "InsureTech Partner 1" but content is a year flag like `2021` for many top accounts.
- **Top Clients Shortlist row 26 has date "1905-07-13"**: Excel epoch artifact.
- **Strategic-status inconsistency**: e.g., Travelers is "Named" in Top Insurers and "Named Backup" in Competitor Analysis.
- **F3 + F4 vintages estimated**: no explicit date in those files; verify with file owners.

## Output files in pipeline/data/

CSVs above + `rows_log.csv` (audit log) + `aliases.csv` + `unresolved.csv`.
Run `audit.py` for the row-by-row reconciliation report.
"""

with open(REPORT, 'w') as f:
    f.write(report_md)

print(f"Done. Status counts: {dict(status_counter)}")
print(f"Total rows logged: {len(rows_log)}")
print(f"Report: {REPORT}")

"""
White-Space Opportunity Map
============================
A one-shot Excel view of where EXL can hunt, organized along three axes:
  - Clients (rows in master heatmap)
  - Insurance value-chain functions (cols)
  - Insurer types / LOB segments (group views)

Sheets generated:
  1. README                          - how to read the file
  2. Master Heatmap                  - Client x Function with competitor footprint
  3. Insurer-Segment Summary         - Insurer Type x Function with counts + EXL solutions
  4. White Space Ranking             - per-client ranking of open functions
  5. Solution x Segment Fit          - which solutions fit which insurer types
  6. Capability Gaps                 - where EXL needs partners (from F1 Cap Gap Ver2)

Output: pipeline/WHITESPACE_MAP.xlsx
"""

import pandas as pd
import os, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(__file__))
from exl_taxonomy import (
    PC_FUNCTIONS, LA_FUNCTIONS, RETENTION_BENCHMARKS,
    FUNCTION_TO_PRODUCTS, EXL_PRODUCTS, retention_headroom,
)

from pathlib import Path
REPO = Path(__file__).resolve().parents[2]                      # repo root (code/pipeline/<script>.py)
DATA = str(REPO / "produced_data" / "pipeline" / "data")
OUT  = str(REPO / "produced_data" / "pipeline" / "WHITESPACE_MAP.xlsx")

# ============ Load ============
profiles = pd.read_csv(f"{DATA}/client_profiles.csv")
ca_raw   = pd.read_csv(f"{DATA}/engagements_competitor_analysis.csv")
ca       = ca_raw[ca_raw['source_file'] == 'F5'].copy()  # F5 canonical
f4       = pd.read_csv(f"{DATA}/engagements_f4_clientwise.csv")
priorities = pd.read_csv(f"{DATA}/priorities.csv")
solutions  = pd.read_csv(f"{DATA}/solutions.csv")
cap_gaps   = pd.read_csv(f"{DATA}/capability_gaps.csv")
ranked     = pd.read_csv(f"{DATA}/leads_ranked.csv")
eng_summary = pd.read_csv(f"{DATA}/client_engagements_summary.csv")

# ============ Function taxonomy ============
# Now using EXL's official 5-area P&C value chain (deck slide 6)
# Plus the typical retention benchmark per area (worst-case = best headroom for sales)
FUNC_NAMES = [
    "Broker Management",
    "Actuarial / NB / UW",
    "Policy Administration",
    "Claims Management",
    "Accounting / Billing",
]

# Map each area to a representative retention benchmark from the Partnering Matrix
# Used for the column header annotations and headroom hints
AREA_TO_BENCHMARK_FN = {
    "Broker Management":     None,  # not in the matrix - data sparse
    "Actuarial / NB / UW":   "Underwriting Risk Assessment",   # 10-20% retention - highest headroom
    "Policy Administration": "Policy Servicing",                # 30-40% retention
    "Claims Management":     "Claims Set up (FNOL)",            # 10-20% retention - highest headroom
    "Accounting / Billing":  None,  # not in matrix
}

# ============ Insurer-type classification ============

# Known L&A clients (from F4 L&A Clientwise canonicalization)
L_A_CLIENTS = set(f4[f4['lob_segment'] == 'L&A']['client_canonical'].dropna())
PC_CLIENTS  = set(f4[f4['lob_segment'] == 'P&C']['client_canonical'].dropna())

# Known reinsurers (name heuristic)
REINSURER_HINTS = ["swiss re", "munich re", "munich-amer", "hannover re", "scor",
                   "renaissancere", "fortitude re", "general re", "berkshire hathaway",
                   "everest re", "transamerica re", "rga re", "partner re", "axis"]

def classify_insurer_type(profile):
    name = str(profile.get('client_canonical', '')).lower()
    typ  = str(profile.get('type', '')).lower()
    canon = profile.get('client_canonical')

    if canon in L_A_CLIENTS:
        return "L&A"
    if "broker" in typ or canon in {"AON", "Marsh", "MMC", "Mercer"}:
        return "Broker"
    if any(h in name for h in REINSURER_HINTS):
        return "Reinsurer"
    if canon in PC_CLIENTS:
        return "P&C"
    if "carrier" in typ:
        return "P&C (presumed)"
    return "Other / Unclassified"

profiles['insurer_type'] = profiles.apply(classify_insurer_type, axis=1)

# ============ Build per-client per-function competitor presence ============

def _entry(r):
    return {
        'competitor': r['competitor_canonical'],
        'ftes': r.get('ftes'),
        'comments': r.get('comments'),
        'renewal': r.get('contract_renewal_window'),
        'source': f"{r['source_file']}/{r['source_sheet']}/R{r['source_row']}",
    }

def func_presence(client, ca_eng):
    """Classifies each engagement into one or more of EXL's 5 official P&C areas
    using flag + keyword heuristics."""
    out = {f: [] for f in FUNC_NAMES}
    for _, r in ca_eng.iterrows():
        c = str(r.get('comments', '')).lower()

        # --- Claims Management ---
        if r.get('claims_y') == 'Y' or any(k in c for k in
                ["claim", "fnol", "bill review", "utilization review", "subrogation", "adjudication"]):
            out["Claims Management"].append(_entry(r))

        # --- Policy Administration (incl Premium Audit) ---
        is_policy_admin = (
            r.get('premium_audit_y') == 'Y'
            or (r.get('uw_y') == 'Y' and any(k in c for k in
                    ["policy admin", "endorsement", "policy servicing", "policy maintenance",
                     "renewal", "cancellation", "coverage verification", "indexing", "mailroom"]))
            or (r.get('platform_y') == 'Y' and any(k in c for k in ["policy", "endorsement"]))
        )
        if is_policy_admin:
            out["Policy Administration"].append(_entry(r))

        # --- Actuarial / NB / UW (UW work that is NOT policy admin) ---
        if r.get('uw_y') == 'Y' and not is_policy_admin:
            out["Actuarial / NB / UW"].append(_entry(r))

        # --- Accounting / Billing ---
        if r.get('fna_y') == 'Y' or any(k in c for k in
                ["collections", "accounts payable", "premium administration", "reconciliation",
                 "billing", "invoic"]):
            out["Accounting / Billing"].append(_entry(r))

        # --- Broker Management ---
        if any(k in c for k in
                ["broker", "agent setup", "agency", "marketing analytics",
                 "new agent", "agent licensing"]):
            out["Broker Management"].append(_entry(r))

    return out

def parse_ftes(v):
    if v is None: return 0
    s = str(v)
    import re
    m = re.search(r'(\d{2,5})', s)
    return int(m.group(1)) if m else 0

def cell_summary(entries):
    """Compact cell content like 'Cognizant 450; Genpact 60' or '' if open"""
    if not entries: return ""
    by_comp = {}
    for e in entries:
        c = e['competitor']
        ftes = parse_ftes(e['ftes'])
        by_comp.setdefault(c, {'ftes': 0, 'friction': False, 'renewal': []})
        by_comp[c]['ftes'] += ftes
        if e.get('renewal'): by_comp[c]['renewal'].append(e['renewal'])
        if e.get('comments') and any(t in str(e['comments']).lower() for t in
                                      ['champion challenger', 'given notice', 'terminating']):
            by_comp[c]['friction'] = True
    parts = []
    for c, d in sorted(by_comp.items(), key=lambda x: -x[1]['ftes']):
        s = c
        if d['ftes']: s += f" ({d['ftes']})"
        if d['friction']: s += "*"
        parts.append(s)
    return "; ".join(parts)

def cell_intensity(entries):
    """Returns intensity 0-4 for color: 0=open, 1=small, 2=medium, 3=large, 4=large+friction"""
    if not entries: return 0
    total_ftes = sum(parse_ftes(e['ftes']) for e in entries)
    has_friction = any(e.get('comments') and any(t in str(e['comments']).lower() for t in
                                                  ['champion challenger', 'given notice', 'terminating'])
                       for e in entries)
    if has_friction and total_ftes >= 100: return 4
    if total_ftes >= 200: return 3
    if total_ftes >= 50: return 2
    if total_ftes > 0: return 1
    if entries: return 1  # competitor present but FTE unknown
    return 0


# ============ Build the matrix ============

# Sort clients: by lead_score, then by NWP rank
client_order = ranked[['rank', 'client_canonical', 'lead_score']].copy()
client_order = client_order.merge(profiles[['client_canonical', 'insurer_type', 'relationship_type',
                                             'nwp_2019', 'rank_2019', 'exl_client', 'exl_growth_leader']],
                                   on='client_canonical', how='left')
# Pick top clients to display - top 50 by lead score OR top NWP rank (whichever is more useful)
# Use union: top 30 by lead score + top 20 by NWP rank
top_by_score = client_order.nlargest(30, 'lead_score', keep='all')['client_canonical'].tolist()
top_by_rank  = client_order.dropna(subset=['rank_2019']).nsmallest(40, 'rank_2019')['client_canonical'].tolist()
display_clients = list(dict.fromkeys(top_by_score + top_by_rank))[:55]  # de-dup, keep order, cap at 55

# Build matrix data
matrix_rows = []
for client in display_clients:
    prof_row = profiles[profiles['client_canonical'] == client]
    if prof_row.empty: continue
    prof = prof_row.iloc[0]
    ca_eng = ca[ca['client_canonical'] == client]
    fp = func_presence(client, ca_eng)

    row = {
        'client_canonical': client,
        'insurer_type': prof['insurer_type'],
        'rank_2019': prof.get('rank_2019'),
        'nwp_2019': prof.get('nwp_2019'),
        'relationship_type': prof.get('relationship_type'),
        'exl_client': prof.get('exl_client'),
        'lead_score': client_order[client_order['client_canonical'] == client]['lead_score'].iloc[0]
                       if not client_order[client_order['client_canonical'] == client].empty else 0,
        'growth_leader': prof.get('exl_growth_leader'),
    }
    n_covered = 0
    n_open = 0
    for func in FUNC_NAMES:
        s = cell_summary(fp[func])
        row[func] = s
        row[func + '__intensity'] = cell_intensity(fp[func])
        if s: n_covered += 1
        else: n_open += 1
    row['num_functions_with_competitor'] = n_covered
    row['num_functions_open'] = n_open
    matrix_rows.append(row)

# ============ Build the Excel file ============

wb = Workbook()

# Styling
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", start_color="1F4E78")
SECTION_FONT = Font(name="Arial", bold=True, color="000000", size=12)
SECTION_FILL = PatternFill("solid", start_color="D9E2F3")
DATA_FONT = Font(name="Arial", size=10)
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)

# Intensity color scale (white -> dark red)
INTENSITY_COLORS = {
    0: "FFFFFF",  # open / white space
    1: "FCE4D6",  # competitor present, small/unknown
    2: "F4B084",  # competitor 50-200 FTEs
    3: "C65911",  # competitor 200+ FTEs
    4: "8B0000",  # competitor + active friction
}
INTENSITY_TEXT_COLORS = {0: "000000", 1: "000000", 2: "000000", 3: "FFFFFF", 4: "FFFFFF"}

def apply_header_style(cell):
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = BORDER

# ============ Sheet 1: README ============
ws = wb.active
ws.title = "README"

readme_content = [
    ("White-Space Opportunity Map", "title"),
    ("", "blank"),
    ("Generated from the consolidated TAM pipeline. Re-generate with: python3 code/pipeline/whitespace.py", "para"),
    ("", "blank"),
    ("How to read this file", "section"),
    ("", "blank"),
    ("1. Master Heatmap — the killer view", "subsection"),
    ("Rows = clients (top by lead score + top by NWP rank). Columns = insurance value-chain functions.", "para"),
    ("Cell content shows competitor presence at that client × function: 'Cognizant (450); Genpact (60)' means", "para"),
    ("Cognizant has ~450 FTEs and Genpact ~60 FTEs in that function. An asterisk (*) means friction signal present.", "para"),
    ("Blank cells = white space (no known competitor in that function at that client).", "para"),
    ("", "blank"),
    ("Color scale:", "para"),
    ("  WHITE      = open white space — nobody documented in that function", "para"),
    ("  LIGHT PEACH= small competitor (<50 FTEs or unknown scale)", "para"),
    ("  ORANGE     = medium competitor (50-199 FTEs)", "para"),
    ("  DARK ORANGE= large competitor (200+ FTEs) - entrenched", "para"),
    ("  DARK RED   = large competitor + active friction (Champion Challenger / notice given)", "para"),
    ("", "blank"),
    ("2. Insurer-Segment Summary", "subsection"),
    ("Aggregates the Master Heatmap by insurer type (P&C, L&A, Reinsurer, Broker). Shows which segments", "para"),
    ("have the most competitor entrenchment in each function and which EXL solutions best fit each cell.", "para"),
    ("", "blank"),
    ("3. White Space Ranking", "subsection"),
    ("Ranks every account by how many functions are 'open' (no documented competitor). High-rank Strategic", "para"),
    ("accounts with many open functions = greenfield opportunities. Sort/filter as needed.", "para"),
    ("", "blank"),
    ("4. Solution × Segment Fit", "subsection"),
    ("Which EXL packaged solutions fit which insurer types, with client appetite ratings from the F1 catalog.", "para"),
    ("", "blank"),
    ("5. Capability Gaps", "subsection"),
    ("From F1 Capability Gap Assessment -Ver2: where EXL has Low/Medium capability and needs a partner", "para"),
    ("(InsureTech / TPA / Other). The 'partner-needed' columns indicate which kind.", "para"),
    ("", "blank"),
    ("Caveats", "section"),
    ("", "blank"),
    ("- FTE counts are best-effort extracted from messy free-text. Treat as order-of-magnitude.", "para"),
    ("- 'Function' classification of each engagement uses both the Y/N column flags AND keyword matching", "para"),
    ("  on Comments (e.g., 'Bill Review' -> Claims). Errors of attribution are possible.", "para"),
    ("- Blank cells DO NOT mean 'EXL is there'. They mean 'no competitor documented in the corpus.'", "para"),
    ("- The view skews to the 30+ accounts with competitor data captured in F1/F4/F5. Many top-NWP accounts", "para"),
    ("  have no engagement data, so their entire row may be blank — that's both an opportunity AND a data gap.", "para"),
    ("- File vintage: F1=Aug 2020, F4=2024+, F5=Feb 2021. Some engagements may be stale.", "para"),
]

r = 1
for text, kind in readme_content:
    cell = ws.cell(row=r, column=1, value=text)
    if kind == "title":
        cell.font = Font(name="Arial", bold=True, size=18, color="1F4E78")
    elif kind == "section":
        cell.font = Font(name="Arial", bold=True, size=14, color="1F4E78")
    elif kind == "subsection":
        cell.font = Font(name="Arial", bold=True, size=11)
    elif kind == "para":
        cell.font = Font(name="Arial", size=10)
    r += 1

ws.column_dimensions['A'].width = 110


# ============ Sheet 2: Master Heatmap ============
ws = wb.create_sheet("Master Heatmap")

# Header rows
ws.cell(row=1, column=1, value="Client × Function Heatmap — competitor footprint, * = active friction signal")
ws.cell(row=1, column=1).font = Font(name="Arial", bold=True, size=14, color="1F4E78")
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(FUNC_NAMES) + 7)

ws.cell(row=2, column=1, value=f"Showing {len(matrix_rows)} clients (top by lead score + top by NWP rank).  Blank = white space.")
ws.cell(row=2, column=1).font = Font(name="Arial", italic=True, size=10, color="666666")
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(FUNC_NAMES) + 7)

# Column headers (row 3 with retention annotation + row 4 with name)
# Row 3: retention/headroom hint per column
hint_row = ["", "", "", "", "", ""]
for fn in FUNC_NAMES:
    bm_label = AREA_TO_BENCHMARK_FN.get(fn)
    if bm_label:
        hr_pct, hr_label = retention_headroom(bm_label)
        if hr_pct is not None:
            hint_row.append(f"Typical EXL retention {100-hr_pct:.0f}% ({hr_label})")
            continue
    hint_row.append("")
hint_row.extend(["", ""])
for c, h in enumerate(hint_row, 1):
    cell = ws.cell(row=3, column=c, value=h)
    cell.font = Font(name="Arial", italic=True, size=9, color="666666")
    cell.alignment = CENTER

# Column headers (row 4)
headers = ["Client", "Insurer Type", "Rel", "EXL Client", "Lead Score", "Rank"] + FUNC_NAMES + ["# Open", "# Covered"]
for c, h in enumerate(headers, 1):
    cell = ws.cell(row=4, column=c, value=h)
    apply_header_style(cell)

# Data rows
for i, mr in enumerate(matrix_rows, 5):
    ws.cell(row=i, column=1, value=mr['client_canonical']).font = DATA_FONT
    ws.cell(row=i, column=2, value=mr['insurer_type']).font = DATA_FONT
    ws.cell(row=i, column=3, value=mr['relationship_type']).font = DATA_FONT
    ws.cell(row=i, column=4, value=mr['exl_client']).font = DATA_FONT
    score_cell = ws.cell(row=i, column=5, value=mr['lead_score'])
    score_cell.font = DATA_FONT
    score_cell.alignment = CENTER
    score_cell.number_format = '0.0'
    rank_cell = ws.cell(row=i, column=6, value=mr.get('rank_2019') if pd.notna(mr.get('rank_2019')) else "")
    rank_cell.font = DATA_FONT
    rank_cell.alignment = CENTER
    if isinstance(rank_cell.value, (int, float)):
        rank_cell.number_format = '0'

    # Function cells with intensity coloring
    for j, func in enumerate(FUNC_NAMES):
        col = 7 + j
        v = mr[func]
        intensity = mr[func + '__intensity']
        cell = ws.cell(row=i, column=col, value=v)
        cell.font = Font(name="Arial", size=9, color=INTENSITY_TEXT_COLORS[intensity])
        cell.fill = PatternFill("solid", start_color=INTENSITY_COLORS[intensity])
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = BORDER

    # Totals
    ws.cell(row=i, column=7 + len(FUNC_NAMES), value=mr['num_functions_open']).alignment = CENTER
    ws.cell(row=i, column=8 + len(FUNC_NAMES), value=mr['num_functions_with_competitor']).alignment = CENTER

    # Score-based row tinting on client column
    s = mr['lead_score']
    if s >= 70:
        ws.cell(row=i, column=1).fill = PatternFill("solid", start_color="00B050")
        ws.cell(row=i, column=1).font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    elif s >= 55:
        ws.cell(row=i, column=1).fill = PatternFill("solid", start_color="92D050")
    elif s >= 40:
        ws.cell(row=i, column=1).fill = PatternFill("solid", start_color="FFC000")
    elif s >= 25:
        ws.cell(row=i, column=1).fill = PatternFill("solid", start_color="FFE699")

# Widths
col_widths = {'A': 30, 'B': 16, 'C': 14, 'D': 11, 'E': 11, 'F': 8}
for col, w in col_widths.items():
    ws.column_dimensions[col].width = w
for j in range(len(FUNC_NAMES)):
    ws.column_dimensions[get_column_letter(7 + j)].width = 26
ws.column_dimensions[get_column_letter(7 + len(FUNC_NAMES))].width = 9
ws.column_dimensions[get_column_letter(8 + len(FUNC_NAMES))].width = 11

ws.freeze_panes = "G5"
ws.auto_filter.ref = f"A4:{get_column_letter(8 + len(FUNC_NAMES))}{4 + len(matrix_rows)}"


# ============ Sheet 3: Insurer-Segment Summary ============
ws = wb.create_sheet("Segment Summary")

# Build per-segment per-function aggregates
seg_counter = {}  # (segment, function) -> {'clients_with_comp': set, 'total_ftes': int, 'top_competitors': Counter}
from collections import Counter
for mr in matrix_rows:
    seg = mr['insurer_type']
    for func in FUNC_NAMES:
        key = (seg, func)
        seg_counter.setdefault(key, {'clients_with_comp': set(), 'total_ftes': 0, 'competitors': Counter()})
        s = mr[func]
        if s:
            seg_counter[key]['clients_with_comp'].add(mr['client_canonical'])
            # Parse competitor names from cell text
            for part in str(s).split(';'):
                part = part.strip().rstrip('*').rstrip(')')
                if '(' in part:
                    name, ftes_str = part.split('(', 1)
                    name = name.strip()
                    try:
                        ftes = int(ftes_str.strip())
                        seg_counter[key]['total_ftes'] += ftes
                    except (ValueError, IndexError):
                        pass
                    seg_counter[key]['competitors'][name] += 1

# EXL product suggestions per official P&C area (real product names from the deck)
FUNC_TO_EXL_SOLUTIONS = {
    "Broker Management":     "EXL Customer 360, EXL Digital Finance Suite",
    "Actuarial / NB / UW":   "EXL Assist, EXL XTRAKTO.AI, EXL MedConnection",
    "Policy Administration": "EXL XTRAKTO.AI, EXL EXELIA.AI, EXL Paymentor, EXL NerveHub",
    "Claims Management":     "EXL XTRAKTO.AI, EXL Subrosource, EXL MedConnection, EXL EXELIA.AI",
    "Accounting / Billing":  "EXL Paymentor, EXL Digital Finance Suite",
}

segments = sorted(set(mr['insurer_type'] for mr in matrix_rows))

ws.cell(row=1, column=1, value="Insurer Segment × Function Summary")
ws.cell(row=1, column=1).font = Font(name="Arial", bold=True, size=14, color="1F4E78")

ws.cell(row=2, column=1, value="For each (insurer segment × value-chain function): # clients with competitor presence; dominant competitors; aggregate FTEs; EXL solution fit.")
ws.cell(row=2, column=1).font = Font(name="Arial", italic=True, size=10, color="666666")
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)

headers = ["Insurer Segment", "Function", "# Clients w/ Competitor", "Top Competitors", "Aggregate FTEs", "EXL Solution Fit"]
for c, h in enumerate(headers, 1):
    cell = ws.cell(row=4, column=c, value=h)
    apply_header_style(cell)

r = 5
for seg in segments:
    for func in FUNC_NAMES:
        key = (seg, func)
        if key not in seg_counter or not seg_counter[key]['clients_with_comp']:
            continue
        d = seg_counter[key]
        top_comps = "; ".join(f"{c}({n})" for c, n in d['competitors'].most_common(3))
        ws.cell(row=r, column=1, value=seg)
        ws.cell(row=r, column=2, value=func)
        ws.cell(row=r, column=3, value=len(d['clients_with_comp']))
        ws.cell(row=r, column=4, value=top_comps)
        ws.cell(row=r, column=5, value=d['total_ftes'] if d['total_ftes'] else "")
        ws.cell(row=r, column=6, value=FUNC_TO_EXL_SOLUTIONS.get(func, ""))
        for c in range(1, 7):
            ws.cell(row=r, column=c).font = DATA_FONT
            ws.cell(row=r, column=c).border = BORDER
            ws.cell(row=r, column=c).alignment = LEFT
        r += 1

for col, w in zip("ABCDEF", [18, 18, 11, 38, 14, 38]):
    ws.column_dimensions[col].width = w

ws.freeze_panes = "A5"
ws.auto_filter.ref = f"A4:F{r-1}"


# ============ Sheet 4: White Space Ranking ============
ws = wb.create_sheet("White Space Ranking")

ws.cell(row=1, column=1, value="White Space Ranking — accounts ranked by # of functions open (no documented competitor)")
ws.cell(row=1, column=1).font = Font(name="Arial", bold=True, size=14, color="1F4E78")

ws.cell(row=2, column=1, value="Useful for: greenfield prospecting (high lead score + high # open functions = best). Sort by columns as needed.")
ws.cell(row=2, column=1).font = Font(name="Arial", italic=True, size=10, color="666666")
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=10)

# Use full client list, not just top
all_clients_matrix = []
for _, prof in profiles.iterrows():
    client = prof['client_canonical']
    if not isinstance(client, str): continue
    ca_eng = ca[ca['client_canonical'] == client]
    fp = func_presence(client, ca_eng)
    n_covered = sum(1 for f in FUNC_NAMES if fp[f])
    n_open = len(FUNC_NAMES) - n_covered
    # Get lead score
    ls_row = ranked[ranked['client_canonical'] == client]
    lead_score = ls_row.iloc[0]['lead_score'] if not ls_row.empty else 0
    insurer_type = prof['insurer_type']
    rank_2019 = prof.get('rank_2019')

    # Dominant competitor across all functions
    all_comps = Counter()
    for func in FUNC_NAMES:
        for e in fp[func]:
            all_comps[e['competitor']] += 1
    dom = all_comps.most_common(1)[0][0] if all_comps else ""

    # Open functions list
    open_funcs = [f for f in FUNC_NAMES if not fp[f]]
    covered_funcs = [f for f in FUNC_NAMES if fp[f]]

    all_clients_matrix.append({
        'client': client,
        'insurer_type': insurer_type,
        'rel': prof.get('relationship_type'),
        'rank': rank_2019,
        'exl_client': prof.get('exl_client'),
        'lead_score': lead_score,
        'n_open': n_open,
        'n_covered': n_covered,
        'dominant_competitor': dom,
        'open_functions': "; ".join(open_funcs),
        'covered_functions': "; ".join(covered_funcs),
    })

# Sort: prefer high lead_score AND high n_open (white space at high-value account)
all_clients_matrix.sort(key=lambda r: (-(r['lead_score']), -r['n_open']))

headers = ["Client", "Insurer Type", "Rel", "Rank", "EXL Client", "Lead Score",
           "# Open Functions", "# Covered Functions", "Dominant Competitor",
           "Open Functions", "Covered Functions"]
for c, h in enumerate(headers, 1):
    cell = ws.cell(row=4, column=c, value=h)
    apply_header_style(cell)

# Limit to clients with any signal (lead_score > 0) OR Strategic/Named status
for i, r in enumerate([m for m in all_clients_matrix
                        if m['lead_score'] > 0 or m['rel'] in ('Strategic', 'Named', 'Named Backup')], 5):
    ws.cell(row=i, column=1, value=r['client']).font = DATA_FONT
    ws.cell(row=i, column=2, value=r['insurer_type']).font = DATA_FONT
    ws.cell(row=i, column=3, value=r['rel']).font = DATA_FONT
    rc = ws.cell(row=i, column=4, value=r['rank'] if pd.notna(r['rank']) else "")
    rc.font = DATA_FONT; rc.alignment = CENTER
    if isinstance(rc.value, (int, float)): rc.number_format = '0'
    ws.cell(row=i, column=5, value=r['exl_client']).font = DATA_FONT
    sc = ws.cell(row=i, column=6, value=r['lead_score'])
    sc.font = DATA_FONT; sc.alignment = CENTER; sc.number_format = '0.0'
    nc = ws.cell(row=i, column=7, value=r['n_open'])
    nc.font = DATA_FONT; nc.alignment = CENTER
    # Color n_open: green if 6 (all open at strategic), red if 0 (fully covered)
    if r['n_open'] >= 5:
        nc.fill = PatternFill("solid", start_color="C6EFCE")
    elif r['n_open'] >= 3:
        nc.fill = PatternFill("solid", start_color="FFEB9C")
    elif r['n_open'] >= 1:
        nc.fill = PatternFill("solid", start_color="FFC7CE")
    ws.cell(row=i, column=8, value=r['n_covered']).alignment = CENTER
    ws.cell(row=i, column=9, value=r['dominant_competitor']).font = DATA_FONT
    ws.cell(row=i, column=10, value=r['open_functions']).font = DATA_FONT
    ws.cell(row=i, column=11, value=r['covered_functions']).font = DATA_FONT
    for c in range(1, 12):
        ws.cell(row=i, column=c).border = BORDER

for col, w in zip("ABCDEFGHIJK", [30, 16, 14, 7, 11, 11, 12, 13, 22, 50, 50]):
    ws.column_dimensions[col].width = w

ws.freeze_panes = "A5"
last_row = 4 + sum(1 for m in all_clients_matrix
                    if m['lead_score'] > 0 or m['rel'] in ('Strategic', 'Named', 'Named Backup'))
ws.auto_filter.ref = f"A4:K{last_row}"


# ============ Sheet 5: Solution × Segment Fit ============
ws = wb.create_sheet("Solution Fit")

ws.cell(row=1, column=1, value="EXL Solution × Insurer Segment Fit (from F1 P&C Solutions List + priorities mapping)")
ws.cell(row=1, column=1).font = Font(name="Arial", bold=True, size=14, color="1F4E78")

# Pull unique offerings from solutions.csv
unique_solutions = solutions[['offering']].drop_duplicates()['offering'].dropna().tolist()

headers = ["Solution", "Client Appetite (from catalog)", "Market Appetite", "EXL Capability (L/M/H)",
           "# Priorities Mapped", "Source Row"]
for c, h in enumerate(headers, 1):
    cell = ws.cell(row=4, column=c, value=h)
    apply_header_style(cell)

# For each solution, get the first row from solutions.csv that has appetite info
r = 5
sol_groups = solutions.groupby('offering').first().reset_index()
for _, s in sol_groups.iterrows():
    ws.cell(row=r, column=1, value=s['offering']).font = DATA_FONT
    ws.cell(row=r, column=2, value=str(s.get('client_appetite', ''))[:200] if pd.notna(s.get('client_appetite')) else "").font = DATA_FONT
    ws.cell(row=r, column=3, value=s.get('market_appetite', '')).font = DATA_FONT
    ws.cell(row=r, column=4, value=s.get('exl_capability_lmh', '')).font = DATA_FONT
    # Count priorities mapping to this solution
    sol_key = str(s['offering']).split(' ')[0] if pd.notna(s['offering']) else ''  # rough match (e.g., "FNOL" matches "FNOL-aaS")
    matched_count = priorities['exl_solution'].fillna('').str.contains(sol_key, case=False, na=False).sum()
    ws.cell(row=r, column=5, value=matched_count).alignment = CENTER
    ws.cell(row=r, column=6, value=f"F1/P&C Solutions List/R{int(s.get('source_row', 0))}").font = DATA_FONT
    # Color by appetite
    appetite = str(s.get('client_appetite', '')).lower() + " " + str(s.get('market_appetite', '')).lower()
    if 'high' in appetite:
        ws.cell(row=r, column=3).fill = PatternFill("solid", start_color="C6EFCE")
    for c in range(1, 7):
        ws.cell(row=r, column=c).border = BORDER
        ws.cell(row=r, column=c).alignment = LEFT
    r += 1

for col, w in zip("ABCDEF", [30, 60, 20, 16, 16, 30]):
    ws.column_dimensions[col].width = w

ws.freeze_panes = "A5"
ws.auto_filter.ref = f"A4:F{r-1}"


# ============ Sheet 6: Capability Gaps ============
ws = wb.create_sheet("Capability Gaps")

ws.cell(row=1, column=1, value="EXL Capability Gaps (from F1 Capability Gap Assessment -Ver2)")
ws.cell(row=1, column=1).font = Font(name="Arial", bold=True, size=14, color="1F4E78")

ws.cell(row=2, column=1, value="Where EXL is Low/Medium capability and needs a partner. The Partner-needed columns indicate kind.")
ws.cell(row=2, column=1).font = Font(name="Arial", italic=True, size=10, color="666666")
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=9)

headers = ["Area", "Solution", "Component", "Feature", "EXL Capability", "Scale 1-5",
           "Needs InsureTech", "Needs TPA", "Needs Other"]
for c, h in enumerate(headers, 1):
    cell = ws.cell(row=4, column=c, value=h)
    apply_header_style(cell)

r = 5
for _, g in cap_gaps.iterrows():
    cap = str(g.get('exl_capability', '')).strip().lower()
    if cap not in ('low', 'medium'): continue  # only show gaps
    ws.cell(row=r, column=1, value=g.get('area'))
    ws.cell(row=r, column=2, value=g.get('solution'))
    ws.cell(row=r, column=3, value=g.get('component'))
    ws.cell(row=r, column=4, value=str(g.get('feature', ''))[:200])
    ec = ws.cell(row=r, column=5, value=g.get('exl_capability'))
    ec.alignment = CENTER
    if cap == 'low':
        ec.fill = PatternFill("solid", start_color="FFC7CE")
    elif cap == 'medium':
        ec.fill = PatternFill("solid", start_color="FFEB9C")
    ws.cell(row=r, column=6, value=g.get('exl_capability_scale_1_5')).alignment = CENTER
    ws.cell(row=r, column=7, value=g.get('partner_insurtech')).alignment = CENTER
    ws.cell(row=r, column=8, value=g.get('partner_tpa')).alignment = CENTER
    ws.cell(row=r, column=9, value=g.get('partner_other')).alignment = CENTER
    for c in range(1, 10):
        ws.cell(row=r, column=c).font = DATA_FONT
        ws.cell(row=r, column=c).border = BORDER
    r += 1

for col, w in zip("ABCDEFGHI", [10, 24, 30, 50, 14, 10, 14, 10, 12]):
    ws.column_dimensions[col].width = w

ws.freeze_panes = "A5"
ws.auto_filter.ref = f"A4:I{r-1}"


# ============ Sheet 7: Partnering Matrix (Retention Benchmarks) ============
ws = wb.create_sheet("Partnering Matrix")

ws.cell(row=1, column=1, value="EXL Partnering Matrix — typical retention by function (from Addressable Market Tracker deck, slides 10-11)")
ws.cell(row=1, column=1).font = Font(name="Arial", bold=True, size=14, color="1F4E78")
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)

ws.cell(row=2, column=1, value="Higher headroom = more of that function typically goes to someone other than EXL = larger displacement TAM when a competitor is present.")
ws.cell(row=2, column=1).font = Font(name="Arial", italic=True, size=10, color="666666")
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)

headers = ["Function", "EXL Retention Range", "Avg Retention", "Headroom %", "Headroom Label"]
for c, h in enumerate(headers, 1):
    cell = ws.cell(row=4, column=c, value=h)
    apply_header_style(cell)

# Sort by headroom descending (largest opportunity first)
rows_data = []
for fn, (lo, hi) in RETENTION_BENCHMARKS.items():
    avg = (lo + hi) / 2
    headroom_pct = round((1 - avg) * 100)
    hr_pct, hr_label = retention_headroom(fn)
    rows_data.append((fn, f"{int(lo*100)}–{int(hi*100)}%", f"{int(avg*100)}%", headroom_pct, hr_label))

rows_data.sort(key=lambda r: -r[3])  # sort by headroom desc

for i, (fn, ret_range, avg, hr_pct, hr_label) in enumerate(rows_data, 5):
    ws.cell(row=i, column=1, value=fn).font = DATA_FONT
    ws.cell(row=i, column=2, value=ret_range).alignment = CENTER
    ws.cell(row=i, column=3, value=avg).alignment = CENTER
    hc = ws.cell(row=i, column=4, value=f"{hr_pct}%")
    hc.alignment = CENTER
    # Color by headroom: green = high opportunity, red = mature/low headroom
    if hr_pct >= 70:
        hc.fill = PatternFill("solid", start_color="00B050")
        hc.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    elif hr_pct >= 50:
        hc.fill = PatternFill("solid", start_color="92D050")
        hc.font = Font(name="Arial", bold=True, size=11)
    elif hr_pct >= 30:
        hc.fill = PatternFill("solid", start_color="FFC000")
    else:
        hc.fill = PatternFill("solid", start_color="FFC7CE")
    ws.cell(row=i, column=5, value=hr_label).font = DATA_FONT
    for c in range(1, 6):
        ws.cell(row=i, column=c).border = BORDER
        if c not in (2, 3, 4):
            ws.cell(row=i, column=c).font = DATA_FONT

for col, w in zip("ABCDE", [34, 18, 14, 13, 26]):
    ws.column_dimensions[col].width = w

ws.freeze_panes = "A5"

# Footnote
fn_row = 5 + len(rows_data) + 2
ws.cell(row=fn_row, column=1, value="How to use this view:")
ws.cell(row=fn_row, column=1).font = Font(name="Arial", bold=True, size=11)
ws.cell(row=fn_row + 1, column=1, value=(
    "1. Find a function where a competitor is parked at one of your accounts (see Master Heatmap)."))
ws.cell(row=fn_row + 2, column=1, value=(
    "2. Look up the headroom % for that function here. >50% headroom = strong displacement opportunity beyond just matching the competitor's FTE count."))
ws.cell(row=fn_row + 3, column=1, value=(
    "3. The `function_headroom` trigger in triggers_fired.csv already fires on this pattern — see lead briefs for accounts where it fires."))
for r_extra in range(fn_row, fn_row + 4):
    ws.cell(row=r_extra, column=1).font = Font(name="Arial", size=10)

# ============ Save ============
wb.save(OUT)
print(f"Saved: {OUT}")
print(f"\nSheets:")
for s in wb.sheetnames:
    print(f"  - {s}")
print(f"\nMaster heatmap shows {len(matrix_rows)} clients.")
print(f"White Space Ranking shows {sum(1 for m in all_clients_matrix if m['lead_score']>0 or m['rel'] in ('Strategic','Named','Named Backup'))} accounts.")

"""
Coverage audit: reconciles every row of every sheet of every file
against the rows_log.csv produced by ingest.py.

For each (file, sheet):
  - total_rows_in_source  = ws.max_row
  - rows_accounted_for    = count in rows_log.csv
  - status_breakdown      = how many ingested / header / empty / etc
  - reconciliation        = OK if total == accounted

Also cross-checks: every output CSV row should have a matching (file, sheet, row)
entry in rows_log.csv with status='ingested'. Catches silent drops.

Output: COVERAGE_AUDIT.md
"""

import openpyxl
import os, csv, warnings
from collections import Counter, defaultdict
warnings.filterwarnings('ignore')

BASE = "/Users/bhavya242574/Library/CloudStorage/OneDrive-EXLService.com(I)Pvt.Ltd/Desktop/TAM/excel_data"
DATA_DIR = "/Users/bhavya242574/Library/CloudStorage/OneDrive-EXLService.com(I)Pvt.Ltd/Desktop/TAM/pipeline/data"
AUDIT = "/Users/bhavya242574/Library/CloudStorage/OneDrive-EXLService.com(I)Pvt.Ltd/Desktop/TAM/pipeline/COVERAGE_AUDIT.md"

FILES = {
    "F1": "20200803 PC-Strategy-Solution Development and GoToMarket PlanVer1.5.xlsx",
    "F2": "Captive Operations Details Jun 2021 v3 - Copy.xlsx",
    "F3": "Competitor Analysis PL and CL.xlsx",
    "F4": "EXL Insurance Competitors.xlsx",
    "F5": "P&C Competitor Analysis.xlsx",
}

# ============ Load source row counts ============

workbooks = {alias: openpyxl.load_workbook(os.path.join(BASE, name), data_only=True)
             for alias, name in FILES.items()}

# (file_alias, sheet) -> max_row
source_max_rows = {}
for alias, wb in workbooks.items():
    for sn in wb.sheetnames:
        source_max_rows[(alias, sn)] = wb[sn].max_row

# ============ Load rows_log ============

log_by_sheet = defaultdict(list)  # (file, sheet) -> list of (row, status, reason)
log_rows_set = set()  # (file, sheet, row) - to compare against CSV output

with open(os.path.join(DATA_DIR, "rows_log.csv"), 'r') as f:
    reader = csv.DictReader(f)
    for row in reader:
        file_alias = row["source_file"]
        sheet = row["source_sheet"]
        row_num = int(row["source_row"])
        status = row["status"]
        reason = row["reason"]
        log_by_sheet[(file_alias, sheet)].append((row_num, status, reason))
        log_rows_set.add((file_alias, sheet, row_num))

# ============ Cross-check: every output CSV row appears in log as 'ingested' ============

output_csvs = [f for f in os.listdir(DATA_DIR) if f.endswith(".csv") and f not in ("rows_log.csv", "aliases.csv", "unresolved.csv")]

orphan_rows = []  # rows in CSV outputs that aren't in log as 'ingested'

ingested_in_log = set()
for (f_alias, sheet), entries in log_by_sheet.items():
    for r, s, _ in entries:
        if s == "ingested":
            ingested_in_log.add((f_alias, sheet, r))

csv_output_rows = set()
for csv_file in output_csvs:
    path = os.path.join(DATA_DIR, csv_file)
    with open(path, 'r') as f:
        reader = csv.DictReader(f)
        for row in reader:
            if "source_file" in row and row["source_file"]:
                try:
                    triple = (row["source_file"], row["source_sheet"], int(row["source_row"]))
                    csv_output_rows.add(triple)
                    if triple not in ingested_in_log:
                        orphan_rows.append((csv_file, triple))
                except (ValueError, KeyError):
                    pass

# Inverse: rows logged as 'ingested' but missing from any output CSV
missing_outputs = [t for t in ingested_in_log if t not in csv_output_rows]


# ============ Build per-sheet reconciliation table ============

rows_data = []
for (f_alias, sheet), max_row in sorted(source_max_rows.items()):
    log_entries = log_by_sheet.get((f_alias, sheet), [])
    logged_count = len(log_entries)
    statuses = Counter(s for _, s, _ in log_entries)
    reconciled = "✓" if logged_count == max_row else f"✗ ({max_row} vs {logged_count})"

    rows_data.append({
        "file": f_alias,
        "sheet": sheet,
        "max_row": max_row,
        "logged": logged_count,
        "ingested": statuses.get("ingested", 0),
        "header": statuses.get("header", 0),
        "empty": statuses.get("empty", 0),
        "explicit_skip": statuses.get("explicitly_skipped", 0),
        "non_data": statuses.get("skipped_non_data", 0),
        "reconciled": reconciled,
    })

# ============ Report ============

totals = {
    "max_row": sum(r["max_row"] for r in rows_data),
    "logged":  sum(r["logged"]  for r in rows_data),
    "ingested": sum(r["ingested"] for r in rows_data),
    "header": sum(r["header"] for r in rows_data),
    "empty": sum(r["empty"] for r in rows_data),
    "explicit_skip": sum(r["explicit_skip"] for r in rows_data),
    "non_data": sum(r["non_data"] for r in rows_data),
}

mismatches = [r for r in rows_data if r["reconciled"] != "✓"]

verdict = "PASS ✓" if not mismatches and not orphan_rows and not missing_outputs else "FAIL ✗"

md = f"""# Coverage Audit

Reconciles every row of every sheet of every Excel file against ingestion log.

## Verdict

**{verdict}**

| Check | Result |
|---|---|
| Sheets with mismatched row counts | {len(mismatches)} |
| Output CSV rows missing from ingestion log | {len(orphan_rows)} |
| Ingested-flagged rows missing from any output CSV | {len(missing_outputs)} |

## Totals

| Metric | Count |
|---|---:|
| Total rows across all sheets (sum of max_row) | {totals['max_row']} |
| Total rows logged in rows_log.csv | {totals['logged']} |
| Total rows ingested into output tables | {totals['ingested']} |
| Header rows (intentionally skipped) | {totals['header']} |
| Empty rows | {totals['empty']} |
| Explicitly-skipped rows (sheet-level skip with reason) | {totals['explicit_skip']} |
| Skipped non-data rows (content present but not data shape) | {totals['non_data']} |
| Sum (ingested + header + empty + explicit_skip + non_data) | {totals['ingested'] + totals['header'] + totals['empty'] + totals['explicit_skip'] + totals['non_data']} |

Reconciliation: `max_row` total should equal `logged` total AND the sum of all categories.

## Per-sheet breakdown

| File | Sheet | max_row | logged | ingested | header | empty | explicit_skip | non_data | Reconciled? |
|---|---|---:|---:|---:|---:|---:|---:|---:|:---:|
"""
for r in rows_data:
    md += (f"| {r['file']} | {r['sheet']} | {r['max_row']} | {r['logged']} | "
           f"{r['ingested']} | {r['header']} | {r['empty']} | {r['explicit_skip']} | "
           f"{r['non_data']} | {r['reconciled']} |\n")

if mismatches:
    md += "\n## Mismatches\n\n"
    for r in mismatches:
        md += f"- **{r['file']} / {r['sheet']}**: source has {r['max_row']} rows, log has {r['logged']}\n"

if orphan_rows:
    md += "\n## Orphan rows in output CSVs (not in log)\n\n"
    for csv_file, triple in orphan_rows[:50]:
        md += f"- `{csv_file}` row references {triple} — but not logged as ingested\n"

if missing_outputs:
    md += "\n## Ingested rows missing from output CSVs\n\n"
    for triple in missing_outputs[:50]:
        md += f"- {triple}\n"

# Add 'skipped_non_data' detail per sheet (these are worth scrutinising — content present, not ingested)
non_data_per_sheet = defaultdict(list)
for (f_alias, sheet), entries in log_by_sheet.items():
    for r, s, reason in entries:
        if s == "skipped_non_data":
            non_data_per_sheet[(f_alias, sheet)].append(r)

if non_data_per_sheet:
    md += "\n## `skipped_non_data` rows — worth scrutinising\n\n"
    md += "These rows have content but didn't match the expected data shape. Could be footer notes, sub-tables, or schema mismatches.\n\n"
    for (f_alias, sheet), rows in sorted(non_data_per_sheet.items()):
        md += f"- **{f_alias} / {sheet}** ({len(rows)} rows): {rows[:20]}{'...' if len(rows) > 20 else ''}\n"

# Datetime coverage check
md += "\n## Datetime coverage on output rows\n\n"
dt_check_rows = []
for csv_file in output_csvs:
    path = os.path.join(DATA_DIR, csv_file)
    with open(path, 'r') as f:
        reader = csv.DictReader(f)
        rows = list(reader)
    if not rows: continue
    has_file_vintage = sum(1 for r in rows if r.get("file_vintage"))
    has_captured_date = sum(1 for r in rows if r.get("captured_date"))
    has_ingestion_date = sum(1 for r in rows if r.get("ingestion_date"))
    dt_check_rows.append([
        csv_file, len(rows), has_file_vintage, has_captured_date, has_ingestion_date,
        "✓" if has_file_vintage == has_captured_date == has_ingestion_date == len(rows) else "✗"
    ])

md += "| Table | Rows | with file_vintage | with captured_date | with ingestion_date | Complete? |\n"
md += "|---|---:|---:|---:|---:|:---:|\n"
for row in dt_check_rows:
    md += f"| {row[0]} | {row[1]} | {row[2]} | {row[3]} | {row[4]} | {row[5]} |\n"

with open(AUDIT, 'w') as f:
    f.write(md)

print(f"Verdict: {verdict}")
print(f"  Sheets with mismatched row counts: {len(mismatches)}")
print(f"  Orphan rows in output CSVs (not in log): {len(orphan_rows)}")
print(f"  Ingested-flagged rows missing from output CSVs: {len(missing_outputs)}")
print(f"  Total max_row across all sheets: {totals['max_row']}")
print(f"  Total logged: {totals['logged']}")
print(f"\nReport: {AUDIT}")
